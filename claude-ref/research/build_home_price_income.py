"""Build research/canada_home_price_income.xlsx

Personal analysis file for Jay Zhao-Murray.
Sources:
  - StatCan 11-10-0190-01: median total income, 2024 constant dollars, 1976-2024
  - CREA MLS HPI Not Seasonally Adjusted (A): Composite_Benchmark (national), 2005-2025
  - StatCan 18-10-0004-01 via WDS vector 41690973: CPI all-items NSA monthly (2002=100)
    annual-averaged to produce deflator

Run from the project root:
    py research/build_home_price_income.py
"""

from __future__ import annotations

import sys
import pathlib

# Allow imports from pipeline/
ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from pipeline.fetch.statcan import fetch_vector

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
INCOME_CSV = ROOT / "data/raw/statcan_11100190/11100190.csv"
CREA_XLSX = ROOT / "data/raw/crea_mls_hpi/Not Seasonally Adjusted (A).xlsx"
OUTPUT_XLSX = ROOT / "research/canada_home_price_income.xlsx"

# CPI vector: 18-10-0004-01, Canada, all-items, NSA (2002=100)
CPI_VECTOR_ID = 41690973

# ---------------------------------------------------------------------------
# 1. Load StatCan income (constant 2024 dollars only)
# ---------------------------------------------------------------------------
print("Loading StatCan 11-10-0190-01 income data...")
income_raw = pd.read_csv(INCOME_CSV)

income = income_raw[
    (income_raw["GEO"] == "Canada")
    & (income_raw["Income concept"] == "Median total income")
    & (income_raw["Economic family type"] == "Economic families and persons not in an economic family")
    & (income_raw["UOM"] == "2024 constant dollars")
].copy()

if income.empty:
    raise RuntimeError(
        "Income filter returned no rows. Check UOM / Income concept labels in "
        f"{INCOME_CSV}. Available Income concepts: {income_raw['Income concept'].unique()}"
    )

income = income[["REF_DATE", "VALUE"]].rename(columns={"REF_DATE": "year", "VALUE": "median_income_real_2024"})
income["year"] = income["year"].astype(int)
income = income.sort_values("year").reset_index(drop=True)

print(f"  Income rows: {len(income)}, years {income['year'].min()}-{income['year'].max()}")
print(f"  1976 value (expect ~74000): {income[income['year']==1976]['median_income_real_2024'].values}")

# ---------------------------------------------------------------------------
# 2. Load CREA annual composite benchmark (nominal $)
# ---------------------------------------------------------------------------
print("Loading CREA MLS HPI annual composite benchmark...")
crea_raw = pd.read_excel(CREA_XLSX, sheet_name="AGGREGATE", header=None)

# Row 0 = headers, rows 1+ = data
headers = crea_raw.iloc[0].tolist()
crea_data = crea_raw.iloc[1:].copy()
crea_data.columns = headers
crea_data = crea_data.reset_index(drop=True)

crea = crea_data[["Date", "Composite_Benchmark"]].rename(
    columns={"Date": "year", "Composite_Benchmark": "home_price_nominal"}
)
crea["year"] = crea["year"].astype(int)
crea["home_price_nominal"] = pd.to_numeric(crea["home_price_nominal"], errors="coerce")
crea = crea.sort_values("year").reset_index(drop=True)

print(f"  CREA rows: {len(crea)}, years {crea['year'].min()}-{crea['year'].max()}")
print(f"  2005 value (expect ~243600): {crea[crea['year']==2005]['home_price_nominal'].values}")

# ---------------------------------------------------------------------------
# 3. Fetch CPI monthly from StatCan WDS and compute annual averages
# ---------------------------------------------------------------------------
print("Fetching CPI all-items NSA monthly from StatCan WDS (vector 41690973)...")
cpi_result = fetch_vector(CPI_VECTOR_ID, latest_n=700)
cpi_monthly = cpi_result.data.copy()
cpi_monthly["year"] = pd.to_datetime(cpi_monthly["date"]).dt.year

# Annual average of monthly NSA CPI
cpi_annual = (
    cpi_monthly.groupby("year")["value"]
    .mean()
    .reset_index()
    .rename(columns={"value": "cpi_index"})
)

# Filter to full years only (drop partial current year if fewer than 12 months)
months_per_year = cpi_monthly.groupby("year")["value"].count()
full_years = months_per_year[months_per_year == 12].index
cpi_annual = cpi_annual[cpi_annual["year"].isin(full_years)].copy()

print(f"  CPI annual rows: {len(cpi_annual)}, years {cpi_annual['year'].min()}-{cpi_annual['year'].max()}")
print(f"  CPI release date: {cpi_result.release_date}")

# ---------------------------------------------------------------------------
# 4. Compute deflator (base year = 2024)
# ---------------------------------------------------------------------------
cpi_2024 = cpi_annual.loc[cpi_annual["year"] == 2024, "cpi_index"].values[0]
print(f"  CPI 2024 (base): {cpi_2024:.4f}")

cpi_annual = cpi_annual.copy()
cpi_annual["deflator"] = cpi_2024 / cpi_annual["cpi_index"]

# ---------------------------------------------------------------------------
# 5. Assemble master frame: union of all years 1976-2025
# ---------------------------------------------------------------------------
all_years = list(range(1976, 2026))
df = pd.DataFrame({"year": all_years})

df = df.merge(cpi_annual[["year", "cpi_index", "deflator"]], on="year", how="left")
df = df.merge(income[["year", "median_income_real_2024"]], on="year", how="left")
df = df.merge(crea[["year", "home_price_nominal"]], on="year", how="left")

# Derive real home price (CREA nominal -> 2024 constant dollars)
df["home_price_real_2024"] = df["home_price_nominal"] * df["deflator"]

# Derive nominal income (inverse of deflator on the constant-dollar series)
# real / deflator = nominal => nominal = real / deflator
df["median_income_nominal"] = df["median_income_real_2024"] / df["deflator"]

# Price-to-income ratio (real/real, equivalently nominal/nominal)
df["price_to_income_ratio"] = df["home_price_real_2024"] / df["median_income_real_2024"]

# Final column order
output_cols = [
    "year",
    "home_price_nominal",
    "home_price_real_2024",
    "median_income_nominal",
    "median_income_real_2024",
    "cpi_index",
    "price_to_income_ratio",
]
df = df[output_cols]

# ---------------------------------------------------------------------------
# 6. Sanity checks
# ---------------------------------------------------------------------------
print("\nSanity checks:")

row_2024 = df[df["year"] == 2024].iloc[0]
print(f"  2024: home_price_nominal={row_2024['home_price_nominal']:.0f}, "
      f"home_price_real_2024={row_2024['home_price_real_2024']:.0f} (should match)")
real_nominal_agree = abs(row_2024["home_price_nominal"] - row_2024["home_price_real_2024"]) < 1.0
print(f"  2024 real==nominal: {'PASS' if real_nominal_agree else 'FAIL'}")

row_income_2024 = df[df["year"] == 2024].iloc[0]
real_income_nominal_agree = abs(row_income_2024["median_income_nominal"] - row_income_2024["median_income_real_2024"]) < 1.0
print(f"  2024 income real==nominal: {'PASS' if real_income_nominal_agree else 'FAIL'}")

row_2005 = df[df["year"] == 2005].iloc[0]
print(f"  2005: home_price_nominal={row_2005['home_price_nominal']:.0f} (expect ~243600)")
hp_2005_ok = abs(row_2005["home_price_nominal"] - 243600) < 500
print(f"  2005 home price: {'PASS' if hp_2005_ok else 'WARN'}")

row_1976 = df[df["year"] == 1976].iloc[0]
print(f"  1976: median_income_real_2024={row_1976['median_income_real_2024']:.0f} (expect ~74000)")
inc_1976_ok = abs(row_1976["median_income_real_2024"] - 74000) < 2000
print(f"  1976 income: {'PASS' if inc_1976_ok else 'WARN'}")

# ---------------------------------------------------------------------------
# 7. Write XLSX with two sheets
# ---------------------------------------------------------------------------
print(f"\nWriting {OUTPUT_XLSX} ...")

wb = openpyxl.Workbook()
ws_data = wb.active
ws_data.title = "data"

# Header style
header_font = Font(bold=True)
header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
thin = Side(style="thin")
border = Border(bottom=Side(style="medium"))

# Write headers
headers_out = [
    "year",
    "home_price_nominal",
    "home_price_real_2024",
    "median_income_nominal",
    "median_income_real_2024",
    "cpi_index",
    "price_to_income_ratio",
]
for col_idx, header in enumerate(headers_out, start=1):
    cell = ws_data.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

# Write data rows
for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    ws_data.cell(row=row_idx, column=1, value=row.year)
    # home_price columns: blank before 2005
    for col_idx, val in enumerate(
        [row.home_price_nominal, row.home_price_real_2024], start=2
    ):
        cell = ws_data.cell(row=row_idx, column=col_idx, value=None if pd.isna(val) else round(val, 0))
        if not pd.isna(val):
            cell.number_format = '#,##0'
    # income columns
    for col_idx, val in enumerate(
        [row.median_income_nominal, row.median_income_real_2024], start=4
    ):
        cell = ws_data.cell(row=row_idx, column=col_idx, value=None if pd.isna(val) else round(val, 0))
        if not pd.isna(val):
            cell.number_format = '#,##0'
    # cpi_index
    cpi_cell = ws_data.cell(row=row_idx, column=6, value=None if pd.isna(row.cpi_index) else round(row.cpi_index, 2))
    cpi_cell.number_format = '0.00'
    # price_to_income_ratio
    pti_cell = ws_data.cell(row=row_idx, column=7, value=None if pd.isna(row.price_to_income_ratio) else round(row.price_to_income_ratio, 2))
    pti_cell.number_format = '0.00'

# Auto-width columns
col_widths = [8, 22, 22, 24, 24, 12, 22]
for i, width in enumerate(col_widths, start=1):
    ws_data.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# Freeze header row
ws_data.freeze_panes = "A2"

# ---------------------------------------------------------------------------
# Metadata sheet
# ---------------------------------------------------------------------------
ws_meta = wb.create_sheet("metadata")
meta_rows = [
    ["Field", "Value"],
    ["Title", "Canada Home Price-to-Income Ratio, 1976-2025"],
    ["Prepared by", "Jay Zhao-Murray / Sibley Creek"],
    ["Prepared date", "2026-05-25"],
    [""],
    ["== SERIES =="],
    ["home_price_nominal",
     "CREA MLS HPI Composite Benchmark, national aggregate, NOT seasonally adjusted annual. "
     "Nominal Canadian dollars. Source: CREA MLS HPI, Not Seasonally Adjusted (A).xlsx, AGGREGATE sheet."],
    ["home_price_real_2024",
     "CREA Composite Benchmark deflated to 2024 constant dollars. "
     "Formula: home_price_nominal * (CPI_2024 / CPI_year). "
     "Available 2005-2025 (CREA HPI coverage)."],
    ["median_income_nominal",
     "Derived: median_income_real_2024 / deflator. "
     "StatCan 11-10-0190-01 publishes only in 2024 constant dollars; "
     "nominal is back-derived for symmetry."],
    ["median_income_real_2024",
     "Median total income, Economic families and persons not in an economic family, Canada, 2024 constant dollars. "
     "Source: StatCan Table 11-10-0190-01, 1976-2024."],
    ["cpi_index",
     "All-items CPI, Canada, not seasonally adjusted, annual average of monthly NSA values. "
     "Index base: 2002=100. "
     "Source: StatCan Table 18-10-0004-01, vector 41690973. "
     "Fetched via StatCan WDS API 2026-05-25."],
    ["price_to_income_ratio",
     "home_price_real_2024 / median_income_real_2024. "
     "Using real/real is equivalent to nominal/nominal; real is documented here "
     "to keep the deflator transparent. Available 2005-2024 (overlap of both series)."],
    [""],
    ["== DEFLATOR =="],
    ["Method",
     "deflator[year] = CPI_2024 / CPI[year]. "
     "Applied to convert CREA nominal prices to 2024 real. "
     "Income series is published directly in 2024 constant dollars by StatCan."],
    ["CPI base year", "2024 (annual average of monthly NSA CPI)"],
    [""],
    ["== PRIMARY SOURCES =="],
    ["StatCan 11-10-0190-01",
     "https://www150.statcan.gc.ca/t1/tbl1/en/tv.action?pid=1110019001"],
    ["StatCan 18-10-0004-01",
     "https://www150.statcan.gc.ca/t1/tbl1/en/tv.action?pid=1810000401"],
    ["CREA MLS HPI",
     "https://www.crea.ca/housing-market-stats/mls-home-price-index/hpi-tool/"],
    [""],
    ["== NOTES =="],
    ["Income coverage", "1976-2024 (StatCan T11-10-0190 vintage as of 2026-05-25)"],
    ["CREA coverage", "2005-2025 (HPI launched 2005; MLS HPI benchmark series)"],
    ["CPI coverage",
     f"Annual averages computed from monthly NSA; 2025 is partial-year estimate "
     f"excluded (only full calendar years used for deflator). "
     f"CPI release date per WDS: {cpi_result.release_date}"],
    ["Constant dollar caveat",
     "StatCan 11-10-0190 switched to 2020 constant dollar base in older vintages; "
     "current release uses 2024 constant dollars throughout (confirmed by UOM field in downloaded CSV)."],
]

for r_idx, row_data in enumerate(meta_rows, start=1):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws_meta.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = Font(bold=True)

ws_meta.column_dimensions["A"].width = 28
ws_meta.column_dimensions["B"].width = 80

wb.save(OUTPUT_XLSX)
print(f"Done. File written: {OUTPUT_XLSX}")
print(f"\nFull 2024 row:")
print(df[df["year"] == 2024].to_string(index=False))
print(f"\nFull 2005 row:")
print(df[df["year"] == 2005].to_string(index=False))
print(f"\nFull 1976 row:")
print(df[df["year"] == 1976].to_string(index=False))
