"""Build a two-panel Excel template that mirrors inflation plate 1, chart 1.

Source convention (matches the live site after the 2026-05-19 fix):
  - Y/Y is computed from the NSA level (StatCan Table 18-10-0004-01),
    which is what StatCan headlines and reporters quote.
  - m/m and 3M-AR are computed from the SA level (StatCan Table
    18-10-0006-01), so the short-horizon rates of change aren't
    contaminated by seasonality.

Both raw level series are pulled monthly by the pipeline; this script
seeds the trailing window from data/raw/cpi_all_items_nsa.csv and
data/raw/cpi_all_items_sa.csv, then builds two native Excel charts:

  LEFT  - bar chart, m/m % CPI (from SA level), trailing 24 months.
          Latest bar red.
  RIGHT - line chart, Y/Y % (from NSA level, solid) + 3M-AR % (from SA
          level, dashed), trailing 36 months. Latest Y/Y point red.

Both charts use flat line caps and no bezier smoothing. They render to
identical physical size so the user can select both, group them once,
and copy-paste as a picture into Word.

Run: py scripts/build_inflation_excel_template.py
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font
from openpyxl.utils.units import pixels_to_EMU

ROOT = Path(__file__).resolve().parent.parent
NSA_CSV = ROOT / "data" / "raw" / "cpi_all_items_nsa.csv"
SA_CSV = ROOT / "data" / "raw" / "cpi_all_items_sa.csv"
OUT_DIR = ROOT / "bylines" / "commentaries"
OUT_PATH = OUT_DIR / "Sibley Creek Inflation Chart Template v2.xlsx"

INK = "15171A"
ACCENT = "C8102E"

LEVEL_HISTORY = 60
MOM_WINDOW = 24
YOY_WINDOW = 36
EDITABLE_RUNWAY = 60


def load_level_csv(path: Path) -> list[tuple[date, float]]:
    rows: list[tuple[date, float]] = []
    with path.open("r", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            d = date.fromisoformat(row["date"])
            v = float(row["value"])
            rows.append((d, v))
    rows.sort(key=lambda r: r[0])
    return rows[-LEVEL_HISTORY:]


def build():
    nsa_list = load_level_csv(NSA_CSV)
    sa_list = load_level_csv(SA_CSV)
    # Build a unified timeline with NSA as the canonical date set. SA
    # usually lags NSA by a month at release (StatCan publishes SA in a
    # separate update), so any tail dates in NSA without an SA value
    # land as blanks in column C — the formula columns return #N/A
    # for those rows and the chart series skip them cleanly.
    sa_by_date = {d: v for d, v in sa_list}
    rows: list[tuple[date, float, float | None]] = []
    for d, n in nsa_list:
        rows.append((d, n, sa_by_date.get(d)))
    # Row 61 will always be the last NSA observation.

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = [
        "Date",
        "NSA Level",
        "SA Level",
        "Y/Y (%)",
        "m/m (%)",
        "3M AR (%)",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(name="Calibri", size=11, bold=True, color=INK)
        c.alignment = Alignment(horizontal="left")

    last_data_row = 1 + len(rows)
    last_sa_row = 1  # tracks the most recent row that has an SA value
    for i, (d, n, s) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=d).number_format = "mmm yyyy"
        ws.cell(row=i, column=2, value=round(n, 3)).number_format = "0.000"
        if s is not None:
            ws.cell(row=i, column=3, value=round(s, 3)).number_format = "0.000"
            last_sa_row = i

    total_rows = LEVEL_HISTORY + EDITABLE_RUNWAY
    for i in range(2, total_rows + 2):
        # Y/Y from NSA: 12-month lag. NA() so chart skips empty rows
        # instead of plotting them as zero.
        if i >= 14:
            ws.cell(
                row=i,
                column=4,
                value=f"=IF(OR(B{i}=\"\",B{i - 12}=\"\"),NA(),(B{i}/B{i - 12}-1)*100)",
            ).number_format = "0.00"
        # m/m from SA: 1-month lag.
        if i >= 3:
            ws.cell(
                row=i,
                column=5,
                value=f"=IF(OR(C{i}=\"\",C{i - 1}=\"\"),NA(),(C{i}/C{i - 1}-1)*100)",
            ).number_format = "0.00"
        # 3M AR from SA: 3-month lag.
        if i >= 5:
            ws.cell(
                row=i,
                column=6,
                value=f"=IF(OR(C{i}=\"\",C{i - 3}=\"\"),NA(),((C{i}/C{i - 3})^4-1)*100)",
            ).number_format = "0.00"

    ws.column_dimensions["A"].width = 13
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 13

    note_row = total_rows + 4
    notes = [
        ("Source — NSA level: StatCan Table 18-10-0004-01 (vector 41690973)."),
        ("Source — SA level:  StatCan Table 18-10-0006-01 (vector 41690914)."),
        (
            "Why two series: Y/Y is computed from NSA so it matches the headline "
            "StatCan publishes; m/m and 3M AR are computed from SA so they aren't "
            "contaminated by seasonality."
        ),
        (
            "Update monthly: paste new Date in column A, new NSA level in B, new SA "
            "level in C. Columns D-F auto-compute. Charts read the trailing 24 / 36 "
            "months from the seeded block (rows 2-61); to roll the window forward "
            "after extending, right-click each chart -> Select Data -> drag source."
        ),
        (
            "Group the two charts (Ctrl-click both -> Format -> Group -> Group), "
            "then right-click -> Copy. Paste-special into Word as PNG."
        ),
    ]
    for offset, txt in enumerate(notes):
        ws.cell(row=note_row + offset, column=1, value=txt).font = Font(
            name="Calibri", size=9, italic=True, color="6B6F76"
        )

    # -------- charts ---------------------------------------------------------
    # The m/m chart ends at the last row that has an SA value; the Y/Y
    # chart ends at the last NSA row.
    mom_last_row = last_sa_row
    yoy_last_row = last_data_row
    mom_first_row = mom_last_row - MOM_WINDOW + 1
    yoy_first_row = yoy_last_row - YOY_WINDOW + 1

    bar = BarChart()
    bar.type = "col"
    bar.style = 2
    bar.title = "MONTH-OVER-MONTH, %"
    bar.legend = None
    bar.y_axis.majorGridlines = None
    bar.x_axis.majorGridlines = None
    bar.gapWidth = 50

    mom_data = Reference(
        ws, min_col=5, min_row=mom_first_row, max_col=5, max_row=mom_last_row
    )
    mom_cats = Reference(
        ws, min_col=1, min_row=mom_first_row, max_row=mom_last_row
    )
    bar.add_data(mom_data, titles_from_data=False)
    bar.set_categories(mom_cats)

    mom_series = bar.series[0]
    mom_series.graphicalProperties = GraphicalProperties(solidFill=INK)
    mom_series.graphicalProperties.line = LineProperties(solidFill=INK)
    latest_idx = MOM_WINDOW - 1
    latest_pt = DataPoint(idx=latest_idx)
    latest_pt.graphicalProperties = GraphicalProperties(solidFill=ACCENT)
    latest_pt.graphicalProperties.line = LineProperties(solidFill=ACCENT)
    mom_series.dPt = [latest_pt]

    _style_axes(bar)

    line = LineChart()
    line.title = "YEAR-OVER-YEAR + 3M AR, %"
    line.legend = None
    line.y_axis.majorGridlines = None
    line.x_axis.majorGridlines = None

    yoy_ref = Reference(
        ws, min_col=4, min_row=yoy_first_row, max_col=4, max_row=yoy_last_row
    )
    ar_ref = Reference(
        ws, min_col=6, min_row=yoy_first_row, max_col=6, max_row=yoy_last_row
    )
    yoy_cats = Reference(
        ws, min_col=1, min_row=yoy_first_row, max_row=yoy_last_row
    )
    line.add_data(yoy_ref, titles_from_data=False)
    line.add_data(ar_ref, titles_from_data=False)
    line.set_categories(yoy_cats)

    yoy_series = line.series[0]
    yoy_series.smooth = False
    yoy_series.graphicalProperties = GraphicalProperties()
    yoy_series.graphicalProperties.line = LineProperties(
        solidFill=INK, w=22000, cap="flat"
    )
    yoy_series.marker = Marker(symbol="none")
    last_yoy_pt = DataPoint(idx=YOY_WINDOW - 1)
    last_yoy_marker = Marker(symbol="circle", size=7)
    last_yoy_marker.graphicalProperties = GraphicalProperties(solidFill=ACCENT)
    last_yoy_marker.graphicalProperties.line = LineProperties(solidFill=ACCENT)
    last_yoy_pt.marker = last_yoy_marker
    yoy_series.dPt = [last_yoy_pt]

    ar_series = line.series[1]
    ar_series.smooth = False
    ar_series.graphicalProperties = GraphicalProperties()
    ar_series.graphicalProperties.line = LineProperties(
        solidFill=INK, w=12700, prstDash="dash", cap="flat"
    )
    ar_series.marker = Marker(symbol="none")

    _style_axes(line)

    panel_w_px = 420
    panel_h_px = 290

    def _anchor(col_pixels: int):
        marker = AnchorMarker(col=7, colOff=pixels_to_EMU(col_pixels),
                              row=0, rowOff=pixels_to_EMU(8))
        ext = XDRPositiveSize2D(cx=pixels_to_EMU(panel_w_px),
                                cy=pixels_to_EMU(panel_h_px))
        return OneCellAnchor(_from=marker, ext=ext)

    bar.anchor = _anchor(0)
    line.anchor = _anchor(panel_w_px + 16)

    ws.add_chart(bar)
    ws.add_chart(line)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)

    latest_nsa = nsa_list[-1]
    latest_sa = sa_list[-1]
    print(f"Wrote {OUT_PATH.relative_to(ROOT)}")
    print(f"  Window: {nsa_list[0][0]} -> {latest_nsa[0]} ({len(nsa_list)} NSA rows)")
    print(f"  Latest NSA level: {latest_nsa[1]:.3f} on {latest_nsa[0]} (drives Y/Y)")
    print(f"  Latest SA  level: {latest_sa[1]:.3f} on {latest_sa[0]} (drives m/m, 3M AR)")
    print(f"  Left chart  (m/m bars):    rows {mom_first_row}-{mom_last_row}")
    print(f"  Right chart (Y/Y + 3M AR): rows {yoy_first_row}-{yoy_last_row}")


def _style_axes(chart) -> None:
    for axis in (chart.x_axis, chart.y_axis):
        axis.delete = False
        axis.spPr = GraphicalProperties()
        axis.spPr.line = LineProperties(solidFill=INK, w=9525)
    chart.y_axis.number_format = '0"%"'


if __name__ == "__main__":
    build()
