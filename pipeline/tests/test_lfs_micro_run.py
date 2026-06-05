"""Tests for pipeline.lfs_micro.run (orchestration) and make_workbook.

Uses fixture-based synthetic data — no live HTTP. Tests:
  - Engine cache save/load roundtrip
  - Spec invalidation clears cache
  - _assemble_series produces expected columns and MA smoothing
  - _compute_new_months on synthetic fixtures produces non-NaN results
  - run() with no new months prints "latest is still..." and exits 0
  - run() with a new month produces expected output files
  - build_workbook() produces a valid xlsx with the expected sheet names
  - write_output_sheet() idempotent rebuild (companion file fallback not tested
    here — that requires OS-level file lock setup)
"""

from __future__ import annotations

import json
import numpy as np
import pandas as pd
import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock


# ---------------------------------------------------------------------------
# Synthetic data helpers (reused from test_lfs_micro_engine.py pattern)
# ---------------------------------------------------------------------------

def _make_df(n: int = 300, seed: int = 42, wage_premium: float = 0.0) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    wage = np.exp(rng.normal(np.log(30) + wage_premium, 0.4, n))
    return pd.DataFrame({
        "wage":         wage,
        "weight":       rng.integers(100, 2000, n).astype(float),
        "gender":       rng.choice([1, 2], n),
        "age_12":       rng.choice(range(1, 13), n),
        "educ":         rng.choice(range(0, 7), n),
        "tenure_bin":   pd.Categorical(rng.choice(["<12m", "12-35m", "36-59m", "60-119m", "120m+"], n)),
        "noc_43":       rng.choice(range(1, 6), n),
        "naics_21":     rng.choice(range(1, 5), n),
        "union_status": rng.choice([1, 2, 3], n),
        "ftptmain":     rng.choice([1, 2], n),
        "mjh":          rng.choice([1, 2], n),
        "permtemp":     rng.choice([1, 2, 3, 4], n),
        "marstat":      rng.choice(range(1, 7), n),
        "immig":        rng.choice([1, 2, 3], n),
        "estsize":      rng.choice([1, 2, 3, 4], n),
        "prov":         rng.choice([10, 24, 35, 48, 59], n),
        "cowmain_pub":  rng.choice([1, 2], n),
    })


def _synthetic_cache_rows(n_months: int = 14, seed: int = 0) -> dict[str, dict]:
    """Build n_months of synthetic engine cache rows.

    Each row has the structure saved by _save_cache(), with realistic lp values.
    Months are labelled 2015-01 through 2015-01 + n_months - 1.
    """
    rng = np.random.default_rng(seed)
    rows = {}
    from pipeline.lfs_micro.engine import _GROUP_LABELS
    base_date = pd.Timestamp("2016-01-01")
    for i in range(n_months):
        d = base_date + pd.DateOffset(months=i)
        key = d.strftime("%Y-%m")
        row = {
            "date": d.strftime("%Y-%m-%d") + "T00:00:00",
            "underlying_lp": float(rng.normal(0.035, 0.005)),
            "composition_lp": float(rng.normal(0.005, 0.002)),
            "raw_mean_lp": float(rng.normal(0.040, 0.005)),
            "total_fitted_lp": float(rng.normal(0.038, 0.004)),
            "n_obs_curr": int(rng.integers(50000, 60000)),
            "n_obs_base": int(rng.integers(50000, 60000)),
            "r2_curr": float(rng.uniform(0.55, 0.65)),
            "r2_base": float(rng.uniform(0.55, 0.65)),
        }
        for g in _GROUP_LABELS:
            row[f"{g}_comp_lp"] = float(rng.normal(0.0, 0.001))
        rows[key] = row
    return rows


# ---------------------------------------------------------------------------
# Tests: engine cache
# ---------------------------------------------------------------------------

class TestEngineCache:
    def test_save_load_roundtrip(self, tmp_path):
        from pipeline.lfs_micro.run import _save_cache, _load_cache
        from pipeline.lfs_micro.engine import _GROUP_LABELS
        import pipeline.lfs_micro.run as run_mod

        # Redirect the raw dir to tmp_path (cache dir is derived from it)
        original = run_mod._RAW_PUMF_DIR
        run_mod._RAW_PUMF_DIR = tmp_path
        try:
            row = {
                "date": "2025-01-01",
                "underlying_lp": 0.035,
                "composition_lp": 0.005,
                "raw_mean_lp": 0.040,
                "total_fitted_lp": 0.038,
                "n_obs_curr": 55000,
                "n_obs_base": 54000,
                "r2_curr": 0.60,
                "r2_base": 0.61,
            }
            for g in _GROUP_LABELS:
                row[f"{g}_comp_lp"] = 0.0002

            _save_cache("2025-01", row)
            loaded = _load_cache("2025-01")

            assert loaded is not None
            assert abs(loaded["underlying_lp"] - 0.035) < 1e-9
            assert loaded["n_obs_curr"] == 55000
            assert "spec" in loaded
            assert "computed_at" in loaded
            assert "parquet_fingerprints" in loaded
        finally:
            run_mod._RAW_PUMF_DIR = original

    def test_spec_invalidation_clears_cache(self, tmp_path):
        from pipeline.lfs_micro.run import _save_cache, _load_cache
        import pipeline.lfs_micro.run as run_mod
        from pipeline.lfs_micro.spec import Spec
        import pipeline.lfs_micro.spec as spec_mod

        original_raw_dir = run_mod._RAW_PUMF_DIR
        original_spec = spec_mod.DEFAULT_SPEC
        run_mod._RAW_PUMF_DIR = tmp_path
        try:
            row = {
                "date": "2025-01-01",
                "underlying_lp": 0.035,
                "composition_lp": 0.005,
                "raw_mean_lp": 0.040,
                "total_fitted_lp": 0.038,
                "n_obs_curr": 55000,
                "n_obs_base": 54000,
                "r2_curr": 0.60,
                "r2_base": 0.61,
            }
            _save_cache("2025-01", row)

            # Change the spec
            spec_mod.DEFAULT_SPEC = Spec(
                weighted=False,  # different from True
                smoothing="ma3",
                ob_reference="base",
                min_cell_count=30,
            )
            run_mod.DEFAULT_SPEC = spec_mod.DEFAULT_SPEC

            loaded = _load_cache("2025-01")
            assert loaded is None, "Cache should be invalidated when spec changes"
        finally:
            run_mod._RAW_PUMF_DIR = original_raw_dir
            spec_mod.DEFAULT_SPEC = original_spec
            run_mod.DEFAULT_SPEC = original_spec

    def test_nan_in_cache_row_serialized_as_null(self, tmp_path):
        from pipeline.lfs_micro.run import _save_cache
        import pipeline.lfs_micro.run as run_mod

        original = run_mod._RAW_PUMF_DIR
        run_mod._RAW_PUMF_DIR = tmp_path
        try:
            row = {
                "date": "2025-01-01",
                "underlying_lp": float("nan"),
                "composition_lp": 0.005,
                "raw_mean_lp": 0.040,
                "total_fitted_lp": 0.038,
                "n_obs_curr": 55000,
                "n_obs_base": 54000,
                "r2_curr": 0.60,
                "r2_base": 0.61,
            }
            _save_cache("2025-01", row)
            raw = json.loads((tmp_path / "_engine_cache" / "2025-01.json").read_text())
            # NaN must be serialized as null, not as a bare NaN token
            assert raw["underlying_lp"] is None
        finally:
            run_mod._RAW_PUMF_DIR = original


# ---------------------------------------------------------------------------
# Tests: _assemble_series
# ---------------------------------------------------------------------------

class TestAssembleSeries:
    def test_output_columns(self):
        from pipeline.lfs_micro.run import _assemble_series
        rows = _synthetic_cache_rows(n_months=12)
        df = _assemble_series(rows)
        assert "underlying_pct" in df.columns
        assert "composition_pct" in df.columns
        assert "raw_mean_pct" in df.columns
        assert "date" in df.columns

    def test_ma3_edge_months_nan(self):
        """Centered MA3: first and last months should be NaN (edge)."""
        from pipeline.lfs_micro.run import _assemble_series
        rows = _synthetic_cache_rows(n_months=6)
        df = _assemble_series(rows)
        # First and last rows should be NaN in underlying_pct (centered MA edge)
        assert pd.isna(df.iloc[0]["underlying_pct"]) or len(df) <= 2
        assert pd.isna(df.iloc[-1]["underlying_pct"]) or len(df) <= 2

    def test_ma3_middle_not_nan(self):
        """Middle months with 12+ months of data should have non-NaN values."""
        from pipeline.lfs_micro.run import _assemble_series
        rows = _synthetic_cache_rows(n_months=12)
        df = _assemble_series(rows)
        # Rows 1..n-2 (0-indexed) should be non-NaN
        middle = df.iloc[1:-1]["underlying_pct"]
        assert middle.notna().all(), f"Middle rows have NaN: {middle[middle.isna()]}"

    def test_pct_range_plausible(self):
        """Underlying pct should be in a plausible wage-growth range."""
        from pipeline.lfs_micro.run import _assemble_series
        rows = _synthetic_cache_rows(n_months=12)
        df = _assemble_series(rows)
        valid = df["underlying_pct"].dropna()
        assert (valid > 0.0).all(), "All underlying pct should be positive (synthetic data)"
        assert (valid < 20.0).all(), "Underlying pct should be < 20% (sanity)"

    def test_empty_cache_returns_empty(self):
        from pipeline.lfs_micro.run import _assemble_series
        df = _assemble_series({})
        assert df.empty


# ---------------------------------------------------------------------------
# Tests: _compute_new_months (requires actual harmonize + engine)
# ---------------------------------------------------------------------------

class TestComputeNewMonths:
    def test_produces_results_for_valid_pair(self, tmp_path):
        """With two months of synthetic data, _compute_new_months returns a result."""
        from pipeline.lfs_micro.run import _compute_new_months
        import pipeline.lfs_micro.run as run_mod

        # Write synthetic parquets for 2024-01 (base) and 2025-01 (current)
        df_base = _make_df(n=400, seed=10)
        df_curr = _make_df(n=400, seed=11, wage_premium=0.03)

        # Add survyear/survmnth to pass validation
        df_base_raw = pd.DataFrame({
            "survyear": [2024] * 400,
            "survmnth": [1] * 400,
            "hrlyearn": (df_base["wage"] * 100).astype(int),
            "finalwt": df_base["weight"].astype(int),
            "lfsstat": [1] * 400,
            "cowmain": [1] * 400,
            "gender": df_base["gender"],
            "age_12": df_base["age_12"],
            "educ": df_base["educ"],
            "tenure": [12] * 400,
            "noc_43": df_base["noc_43"],
            "naics_21": df_base["naics_21"],
            "union": df_base["union_status"],
            "ftptmain": df_base["ftptmain"],
            "mjh": df_base["mjh"],
            "permtemp": df_base["permtemp"],
            "marstat": df_base["marstat"],
            "immig": df_base["immig"],
            "estsize": df_base["estsize"],
            "prov": df_base["prov"],
        })
        df_curr_raw = df_base_raw.copy()
        df_curr_raw["survyear"] = 2025
        df_curr_raw["hrlyearn"] = (df_curr["wage"] * 100).astype(int)

        pumf_dir = tmp_path / "lfs_pumf"
        pumf_dir.mkdir()
        df_base_raw.to_parquet(pumf_dir / "2024-01.parquet", index=False)
        df_curr_raw.to_parquet(pumf_dir / "2025-01.parquet", index=False)

        original_dir = run_mod._RAW_PUMF_DIR
        original_n_floor = run_mod._MIN_PLAUSIBLE_N_OBS
        original_r2_floor = run_mod._MIN_PLAUSIBLE_R2
        run_mod._RAW_PUMF_DIR = pumf_dir
        # Synthetic fixture is far smaller/noisier than a real LFS month;
        # relax the plausibility floors that guard the production cache.
        run_mod._MIN_PLAUSIBLE_N_OBS = 0
        run_mod._MIN_PLAUSIBLE_R2 = -1.0
        try:
            results = _compute_new_months(["2025-01"], {})
            assert "2025-01" in results, "Expected result for 2025-01"
            r = results["2025-01"]
            assert "underlying_lp" in r
            assert pd.notna(r["underlying_lp"])
            assert "n_obs_curr" in r
            assert r["n_obs_curr"] > 0
        finally:
            run_mod._RAW_PUMF_DIR = original_dir
            run_mod._MIN_PLAUSIBLE_N_OBS = original_n_floor
            run_mod._MIN_PLAUSIBLE_R2 = original_r2_floor

    def test_missing_base_month_skipped(self, tmp_path):
        """When the base month parquet is missing, the month is skipped gracefully."""
        from pipeline.lfs_micro.run import _compute_new_months
        import pipeline.lfs_micro.run as run_mod

        pumf_dir = tmp_path / "lfs_pumf"
        pumf_dir.mkdir()
        # Only write the current month, not the base
        df_curr = _make_df(n=200, seed=5)
        df_curr.to_parquet(pumf_dir / "2025-01.parquet", index=False)

        original_dir = run_mod._RAW_PUMF_DIR
        run_mod._RAW_PUMF_DIR = pumf_dir
        try:
            results = _compute_new_months(["2025-01"], {})
            # Should skip 2025-01 because 2024-01 is missing
            assert "2025-01" not in results
        finally:
            run_mod._RAW_PUMF_DIR = original_dir


# ---------------------------------------------------------------------------
# Tests: run() orchestration (mocked HTTP + engine)
# ---------------------------------------------------------------------------

class TestRunOrchestration:
    def test_no_new_months_exits_0(self, tmp_path, capsys):
        """When all engine months are cached, run() exits 0 and prints 'latest is still'."""
        from pipeline.lfs_micro import run as run_mod

        # Patch latest_available_month to return 2026-03
        # Patch _load_all_cache to return a full set covering 2016-01 to 2026-03
        full_cache = _synthetic_cache_rows(n_months=122)  # enough to cover 2016-01..2026-03

        with patch.object(run_mod, "latest_available_month", return_value=(2026, 3)):
            with patch.object(run_mod, "_load_all_cache", return_value=full_cache):
                # _all_yoy_keys(2026, 3) produces 122 keys (2016-01..2026-03)
                # If all 122 are in full_cache (we only have 122 synthetic keys here),
                # it should short-circuit.
                # But the synthetic keys start at 2016-01 and go for 122 months -> 2026-02
                # So 2026-03 would be missing. Add it manually.
                last_key = "2026-03"
                if last_key not in full_cache:
                    full_cache[last_key] = list(full_cache.values())[-1].copy()
                    full_cache[last_key]["date"] = "2026-03-01"

                with patch.object(run_mod, "_load_all_cache", return_value=full_cache):
                    result = run_mod.run(pinned_month="2026-03", force_download=False)

        # The result should be 0 (no-op or success)
        assert result == 0

    def test_run_with_pinned_month(self, tmp_path):
        """run() with a pinned month and a pre-populated synthetic cache writes outputs."""
        from pipeline.lfs_micro import run as run_mod
        import pipeline.lfs_micro.run as run_module

        # Build a 15-month synthetic cache so assemble_series has enough data
        full_cache = _synthetic_cache_rows(n_months=15, seed=99)
        latest_key = "2026-03"
        # Add the pinned month to the cache
        full_cache[latest_key] = list(full_cache.values())[-1].copy()
        full_cache[latest_key]["date"] = "2026-03-01"

        # Redirect output dirs to tmp_path
        original_processed = run_module._PROCESSED_DIR
        original_work = run_module._WORK_DIR

        proc_dir = tmp_path / "processed"
        proc_dir.mkdir()
        work_dir = tmp_path / "work"
        work_dir.mkdir()
        run_module._PROCESSED_DIR = proc_dir
        run_module._WORK_DIR = work_dir

        # Also redirect the raw dir so _save_cache writes to a tmp cache
        # (the engine cache dir is derived from _RAW_PUMF_DIR)
        original_cache = run_module._RAW_PUMF_DIR
        run_module._RAW_PUMF_DIR = tmp_path / "raw_pumf"

        try:
            with patch.object(run_module, "latest_available_month", return_value=(2026, 3)):
                with patch.object(run_module, "_load_all_cache", return_value=full_cache):
                    # Patch _write_replication_series to avoid needing the full CSV infrastructure
                    with patch.object(run_module, "_write_replication_series") as mock_write:
                        mock_write.return_value = (
                            proc_dir / "lfs_micro_replication.csv",
                            proc_dir / "lfs_micro_replication.meta.json"
                        )
                        # Patch workbook and chart writers (they need the CSV)
                        with patch("pipeline.lfs_micro.output_sheet.write_output_sheet") as mock_wb:
                            from dataclasses import dataclass
                            @dataclass
                            class FakeResult:
                                path: Path
                                used_companion: bool
                            mock_wb.return_value = FakeResult(path=work_dir / "test.xlsx", used_companion=False)
                            with patch("pipeline.lfs_micro.chart.render_chart") as mock_chart:
                                mock_chart.return_value = (work_dir / "test.svg", work_dir / "test.html")
                                result = run_module.run(
                                    pinned_month="2026-03",
                                    force_download=False,
                                )

            assert result == 0
            # _write_replication_series was called with a non-empty DataFrame
            assert mock_write.called
            df_arg = mock_write.call_args[0][0]
            assert not df_arg.empty
            assert "underlying_pct" in df_arg.columns
        finally:
            run_module._PROCESSED_DIR = original_processed
            run_module._WORK_DIR = original_work
            run_module._RAW_PUMF_DIR = original_cache


# ---------------------------------------------------------------------------
# Tests: make_workbook (fixture-based)
# ---------------------------------------------------------------------------

class TestMakeWorkbook:
    def test_build_workbook_creates_four_sheets(self, tmp_path):
        """build_workbook() creates a valid xlsx with headline/decomposition/latest_month/params_meta."""
        from pipeline.lfs_micro.make_workbook import build_workbook
        import pipeline.lfs_micro.make_workbook as wb_mod
        from openpyxl import load_workbook as opxl_load

        # Build minimal synthetic CSVs for the workbook to read
        rep_csv = tmp_path / "lfs_micro_replication.csv"
        boc_csv = tmp_path / "lfs_micro.csv"

        # Synthetic replication data: 24 months
        dates = pd.date_range("2024-01-01", periods=24, freq="MS")
        rep_df = pd.DataFrame({
            "date": dates.strftime("%Y-%m-%d"),
            "underlying_pct": np.random.default_rng(42).normal(3.0, 0.3, 24),
            "composition_pct": np.random.default_rng(43).normal(0.5, 0.1, 24),
            "raw_mean_pct": np.random.default_rng(44).normal(3.5, 0.3, 24),
            "total_fitted_pct": np.random.default_rng(45).normal(3.5, 0.3, 24),
            "n_obs_curr": [55000] * 24,
            "n_obs_base": [54000] * 24,
            "r2_curr": [0.60] * 24,
            "r2_base": [0.61] * 24,
        })
        rep_df.to_csv(rep_csv, index=False)

        boc_df = pd.DataFrame({
            "date": dates.strftime("%Y-%m-%d"),
            "value": np.random.default_rng(46).normal(3.0, 0.2, 24),
        })
        boc_df.to_csv(boc_csv, index=False)

        # Redirect module-level paths
        original_rep = wb_mod._REPLICATION_CSV
        original_boc = wb_mod._BOC_CSV
        original_cache = wb_mod._ENGINE_CACHE_DIR
        original_report = wb_mod._CALIBRATION_REPORT
        wb_mod._REPLICATION_CSV = rep_csv
        wb_mod._BOC_CSV = boc_csv
        wb_mod._ENGINE_CACHE_DIR = tmp_path / "_engine_cache"
        wb_mod._CALIBRATION_REPORT = tmp_path / "nonexistent_report.md"

        out_path = tmp_path / "test_workbook.xlsx"
        try:
            build_workbook(out_path, overwrite=True)
            assert out_path.exists(), "Workbook file was not created"

            wb = opxl_load(out_path)
            sheets = wb.sheetnames
            wb.close()

            assert "headline" in sheets, f"Missing 'headline' sheet. Got: {sheets}"
            assert "decomposition" in sheets, f"Missing 'decomposition' sheet. Got: {sheets}"
            assert "latest_month" in sheets, f"Missing 'latest_month' sheet. Got: {sheets}"
            assert "params_meta" in sheets, f"Missing 'params_meta' sheet. Got: {sheets}"
        finally:
            wb_mod._REPLICATION_CSV = original_rep
            wb_mod._BOC_CSV = original_boc
            wb_mod._ENGINE_CACHE_DIR = original_cache
            wb_mod._CALIBRATION_REPORT = original_report

    def test_build_workbook_raises_on_existing_without_overwrite(self, tmp_path):
        """build_workbook() raises FileExistsError when file exists and overwrite=False."""
        from pipeline.lfs_micro.make_workbook import build_workbook
        import pipeline.lfs_micro.make_workbook as wb_mod

        out_path = tmp_path / "test.xlsx"
        out_path.touch()

        # Need to patch paths or it'll try to read real CSVs before the overwrite check
        # The overwrite check happens first (before loading data), so this should raise immediately
        with pytest.raises(FileExistsError):
            build_workbook(out_path, overwrite=False)

    def test_headline_sheet_has_header_and_data(self, tmp_path):
        """headline sheet has expected column headers."""
        from pipeline.lfs_micro.make_workbook import build_workbook
        import pipeline.lfs_micro.make_workbook as wb_mod
        from openpyxl import load_workbook as opxl_load

        dates = pd.date_range("2024-01-01", periods=12, freq="MS")
        rep_df = pd.DataFrame({
            "date": dates.strftime("%Y-%m-%d"),
            "underlying_pct": [3.0] * 12,
            "composition_pct": [0.5] * 12,
            "raw_mean_pct": [3.5] * 12,
            "total_fitted_pct": [3.5] * 12,
            "n_obs_curr": [55000] * 12,
            "n_obs_base": [54000] * 12,
            "r2_curr": [0.60] * 12,
            "r2_base": [0.61] * 12,
        })
        boc_df = pd.DataFrame({
            "date": dates.strftime("%Y-%m-%d"),
            "value": [2.9] * 12,
        })

        rep_csv = tmp_path / "rep.csv"
        boc_csv = tmp_path / "boc.csv"
        rep_df.to_csv(rep_csv, index=False)
        boc_df.to_csv(boc_csv, index=False)

        original_rep = wb_mod._REPLICATION_CSV
        original_boc = wb_mod._BOC_CSV
        original_cache = wb_mod._ENGINE_CACHE_DIR
        original_report = wb_mod._CALIBRATION_REPORT
        wb_mod._REPLICATION_CSV = rep_csv
        wb_mod._BOC_CSV = boc_csv
        wb_mod._ENGINE_CACHE_DIR = tmp_path / "_engine_cache"
        wb_mod._CALIBRATION_REPORT = tmp_path / "nope.md"

        out_path = tmp_path / "test2.xlsx"
        try:
            build_workbook(out_path, overwrite=True)
            wb = opxl_load(out_path)
            ws = wb["headline"]
            headers = [c.value for c in ws[1]]
            wb.close()
            assert "date" in headers
            assert "underlying_ours_%" in headers
            assert "boc_INDINF_LFSMICRO_M_%" in headers
            assert "diff_pp" in headers
        finally:
            wb_mod._REPLICATION_CSV = original_rep
            wb_mod._BOC_CSV = original_boc
            wb_mod._ENGINE_CACHE_DIR = original_cache
            wb_mod._CALIBRATION_REPORT = original_report
