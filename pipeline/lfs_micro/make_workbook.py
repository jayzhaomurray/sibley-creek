"""Build the LFS-micro replication workbook.

Reads from:
  data/processed/lfs_micro_replication.csv  -- composition-adjusted series (ours)
  data/raw/lfs_micro.csv                    -- BoC Valet INDINF_LFSMICRO_M (benchmark)
  data/raw/lfs_pumf/_engine_cache/          -- per-month engine results (group contributions)
  claude-ref/research/lfs_micro/calibration_report.md  -- diagnosis conclusion

Output:
  work/research/lfs_micro/lfs_micro_replication.xlsx

Four sheets:
  headline       -- monthly rows: ours / BoC / diff / raw / composition + line chart
  decomposition  -- per-month group contributions to composition effect
  latest_month   -- newest month detail (decomposition arithmetic, top contributors)
  params_meta    -- frozen Spec, calibration stats, divergence diagnosis, provenance

Styling conventions follow pipeline/shadow_rate/make_workbook.py and output_sheet.py:
  restrained, thin grey rules, dark header bar, frozen panes, no chartjunk.

Windows file-lock: if the xlsx is open in Excel, save to a companion file
  lfs_micro_replication_output.xlsx in the same folder.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parents[2]

DEFAULT_OUT = PROJECT_ROOT / "work" / "research" / "lfs_micro" / "lfs_micro_replication.xlsx"

_REPLICATION_CSV = PROJECT_ROOT / "data" / "processed" / "lfs_micro_replication.csv"
_BOC_CSV = PROJECT_ROOT / "data" / "raw" / "lfs_micro.csv"
_ENGINE_CACHE_DIR = PROJECT_ROOT / "data" / "raw" / "lfs_pumf" / "_engine_cache"
_CALIBRATION_REPORT = PROJECT_ROOT / "claude-ref" / "research" / "lfs_micro" / "calibration_report.md"

# --- styling primitives (matches shadow_rate/output_sheet.py conventions) ---
_INK = "1A1A1A"
_GREY = "BFBFBF"
_HEADER_FILL = PatternFill("solid", fgColor="222222")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri")
_SUBHEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
_NOTE_FONT = Font(italic=True, color="888888", size=9, name="Calibri")
_BODY_FONT = Font(name="Calibri", size=10)
_BOLD_FONT = Font(bold=True, name="Calibri", size=10)
_THIN = Side(style="thin", color=_GREY)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NUM3 = "0.000"
_NUM2 = "0.00"
_NUM1 = "0.0"

_HIGHLIGHT_POS = PatternFill("solid", fgColor="E8F5E9")
_HIGHLIGHT_NEG = PatternFill("solid", fgColor="FFEBEE")

GROUP_LABELS = [
    "occupation", "education", "tenure", "age", "gender", "union",
    "fullparttime", "province", "jobpermanency", "maritalstatus",
    "immigration", "industry", "sector", "estsize", "multijob",
]


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_replication() -> pd.DataFrame:
    df = pd.read_csv(_REPLICATION_CSV, parse_dates=["date"])
    df = df.sort_values("date").reset_index(drop=True)
    return df


def _load_boc() -> pd.Series:
    df = pd.read_csv(_BOC_CSV, parse_dates=["date"])
    df = df.set_index("date").sort_index()
    return df["value"].astype(float)


def _load_engine_cache() -> dict[str, dict]:
    """Load per-month engine cache JSONs. Returns {YYYY-MM: {group: comp_pct, ...}}."""
    results: dict[str, dict] = {}
    if not _ENGINE_CACHE_DIR.exists():
        return results
    for f in sorted(_ENGINE_CACHE_DIR.glob("*.json")):
        key = f.stem  # e.g. "2025-04"
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            results[key] = data
        except Exception:
            pass
    return results


def _calibration_conclusion() -> str:
    """Extract the divergence diagnosis conclusion from the calibration report."""
    if not _CALIBRATION_REPORT.exists():
        return "See calibration report (not found at expected path)."
    text = _CALIBRATION_REPORT.read_text(encoding="utf-8", errors="replace")
    # Find the root cause paragraph
    marker = "### Root cause"
    if marker in text:
        start = text.index(marker)
        # Take the next 1000 chars
        excerpt = text[start:start + 1200].strip()
        # Clean to first double-newline after the substantial content
        lines = excerpt.split("\n")
        out = []
        for line in lines[:15]:
            out.append(line.strip())
        return " ".join(l for l in out if l and not l.startswith("#"))
    return "See calibration_report.md for full diagnosis."


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def _style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _BORDER


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _apply_body_style(ws, min_row: int, max_row: int, ncols: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _BODY_FONT
            cell.border = _BORDER


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_headline_sheet(
    wb: Workbook,
    rep: pd.DataFrame,
    boc: pd.Series,
    last_24_months: bool = True,
) -> None:
    """Sheet 1: headline time series + line chart."""
    ws = wb.active
    ws.title = "headline"

    # Align rep and BoC on common dates
    rep_idx = rep.set_index("date")
    common = rep_idx.index.intersection(boc.index)

    # Sort descending (newest first) — last 24 at top, scrollable
    dates = sorted(common, reverse=True)

    headers = [
        "date",
        "underlying_ours_%",
        "boc_INDINF_LFSMICRO_M_%",
        "diff_pp",
        "raw_mean_wage_yoy_%",
        "composition_effect_%",
        "n_obs",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for d in dates:
        r_row = rep_idx.loc[d] if d in rep_idx.index else None
        boc_v = float(boc.loc[d]) if d in boc.index else None
        if r_row is None:
            continue

        underlying = round(float(r_row["underlying_pct"]), 3) if pd.notna(r_row.get("underlying_pct")) else None
        raw_m = round(float(r_row["raw_mean_pct"]), 3) if "raw_mean_pct" in r_row and pd.notna(r_row.get("raw_mean_pct")) else None
        comp = round(float(r_row["composition_pct"]), 3) if "composition_pct" in r_row and pd.notna(r_row.get("composition_pct")) else None
        diff = round(underlying - boc_v, 3) if (underlying is not None and boc_v is not None) else None
        n_obs = int(r_row["n_obs_curr"]) if "n_obs_curr" in r_row and pd.notna(r_row.get("n_obs_curr")) else None

        ws.append([
            d.strftime("%Y-%m"),
            underlying,
            boc_v,
            diff,
            raw_m,
            comp,
            n_obs,
        ])

    # Format numeric columns
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
        for c in row:
            if c.value is not None:
                c.number_format = _NUM3

    _apply_body_style(ws, 2, ws.max_row, len(headers))

    # Highlight diff column: positive diff = light green, negative = light red
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for c in row:
            if c.value is not None and isinstance(c.value, (int, float)):
                if c.value > 0.2:
                    c.fill = _HIGHLIGHT_POS
                elif c.value < -0.2:
                    c.fill = _HIGHLIGHT_NEG

    _autosize(ws, {1: 10, 2: 18, 3: 24, 4: 10, 5: 20, 6: 18, 7: 10})
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    # --- Native LineChart (ours vs BoC) ---
    # Data runs from row 2 downward (newest first).
    # Only include months where both series have values (first len(common) rows).
    n_data = len(dates)
    if n_data >= 2:
        chart = LineChart()
        chart.title = None
        chart.style = 10
        chart.height = 12
        chart.width = 22

        # Col B: ours; Col C: BoC
        ref_ours = Reference(ws, min_col=2, min_row=1, max_row=n_data + 1)
        ref_boc = Reference(ws, min_col=3, min_row=1, max_row=n_data + 1)
        ref_dates = Reference(ws, min_col=1, min_row=2, max_row=n_data + 1)

        chart.add_data(ref_ours, titles_from_data=True)
        chart.add_data(ref_boc, titles_from_data=True)
        chart.set_categories(ref_dates)

        chart.series[0].graphicalProperties.line.solidFill = "C0392B"
        chart.series[0].graphicalProperties.line.width = 20000
        chart.series[1].graphicalProperties.line.solidFill = "1A1A1A"
        chart.series[1].graphicalProperties.line.width = 14000

        # Place chart below the data
        ws.add_chart(chart, f"A{n_data + 4}")


def _build_decomposition_sheet(
    wb: Workbook,
    rep: pd.DataFrame,
    engine_cache: dict[str, dict],
) -> None:
    """Sheet 2: per-month group contributions to composition effect."""
    ws = wb.create_sheet("decomposition")

    # Build group columns from available cache keys
    # Group contributions in cache are in log-points; convert to pct
    headers = ["date", "composition_effect_%"] + [f"{g}_contrib_%"for g in GROUP_LABELS]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    rep_idx = rep.set_index("date")
    dates = sorted(rep_idx.index, reverse=True)

    for d in dates:
        key = d.strftime("%Y-%m")
        r_row = rep_idx.loc[d]
        comp = round(float(r_row["composition_pct"]), 4) if "composition_pct" in r_row and pd.notna(r_row.get("composition_pct")) else None

        row_vals: list = [d.strftime("%Y-%m"), comp]

        cache = engine_cache.get(key, {})
        for g in GROUP_LABELS:
            lp_val = cache.get(f"{g}_comp_lp")
            if lp_val is not None and not np.isnan(lp_val):
                pct_val = round((np.exp(lp_val) - 1.0) * 100.0, 4)
            else:
                pct_val = None
            row_vals.append(pct_val)

        ws.append(row_vals)

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=len(headers)):
        for c in row:
            if c.value is not None:
                c.number_format = _NUM3

    _apply_body_style(ws, 2, ws.max_row, len(headers))

    # Autosize: date=10, composition=18, each group=14
    widths = {1: 10, 2: 18}
    for i, _ in enumerate(GROUP_LABELS, start=3):
        widths[i] = 14
    _autosize(ws, widths)
    ws.freeze_panes = "A2"

    # Note if cache is empty
    if not engine_cache:
        note_row = ws.max_row + 2
        ws.cell(row=note_row, column=1,
                value="NOTE: Per-month group contribution cache is empty. "
                      "Run the engine with --save-cache to populate this sheet.").font = _NOTE_FONT


def _build_latest_month_sheet(
    wb: Workbook,
    rep: pd.DataFrame,
    boc: pd.Series,
    engine_cache: dict[str, dict],
) -> None:
    """Sheet 3: newest month detail."""
    ws = wb.create_sheet("latest_month")

    rep_idx = rep.set_index("date").sort_index()
    latest_rep = rep_idx.index.max() if not rep_idx.empty else None

    if latest_rep is None:
        ws.cell(1, 1, "No replication data available.").font = _NOTE_FONT
        return

    r_row = rep_idx.loc[latest_rep]
    key = latest_rep.strftime("%Y-%m")
    boc_v = float(boc.loc[latest_rep]) if latest_rep in boc.index else None
    underlying = float(r_row["underlying_pct"]) if pd.notna(r_row.get("underlying_pct")) else None
    raw_m = float(r_row["raw_mean_pct"]) if pd.notna(r_row.get("raw_mean_pct")) else None
    comp = float(r_row["composition_pct"]) if pd.notna(r_row.get("composition_pct")) else None
    diff = round(underlying - boc_v, 3) if (underlying is not None and boc_v is not None) else None
    n_curr = int(r_row["n_obs_curr"]) if pd.notna(r_row.get("n_obs_curr")) else None
    n_base = int(r_row["n_obs_base"]) if pd.notna(r_row.get("n_obs_base")) else None
    r2_curr = float(r_row["r2_curr"]) if pd.notna(r_row.get("r2_curr")) else None

    # --- header ---
    title_cell = ws.cell(1, 1, f"Latest month detail: {key}")
    title_cell.font = Font(bold=True, size=13, name="Calibri")
    ws.merge_cells("A1:D1")

    r = 3

    def _kv(label: str, value, fmt: str | None = None) -> None:
        nonlocal r
        ws.cell(r, 1, label).font = _BOLD_FONT
        c = ws.cell(r, 2, value)
        c.font = _BODY_FONT
        if fmt and value is not None:
            c.number_format = fmt
        r += 1

    _kv("Reference period", key)
    _kv("Sample size (current month)", n_curr, "#,##0")
    _kv("Sample size (base month, t-12)", n_base, "#,##0")
    _kv("R-squared (current month)", r2_curr, _NUM3)
    r += 1

    _kv("Raw mean wage growth (y/y %)", raw_m, _NUM3)
    _kv("Composition effect (%)", comp, _NUM3)
    _kv("Underlying wage growth — ours (y/y %)", underlying, _NUM3)
    _kv("BoC INDINF_LFSMICRO_M (y/y %)", boc_v, _NUM3)
    _kv("Difference (ours minus BoC, pp)", diff, _NUM3)
    r += 1

    # Decomposition arithmetic
    ws.cell(r, 1, "O-B identity check: raw = underlying + composition").font = _BOLD_FONT
    r += 1
    if underlying is not None and comp is not None and raw_m is not None:
        total_ob = round(underlying + comp, 4)
        ws.cell(r, 1, f"  underlying {underlying:.3f} + composition {comp:.3f} = {total_ob:.3f}  (raw: {raw_m:.3f})").font = _NOTE_FONT
        r += 1
    r += 1

    # Group contributions
    cache = engine_cache.get(key, {})
    if cache:
        ws.cell(r, 1, "Group contributions to composition effect").font = _BOLD_FONT
        r += 1

        # Header row for group table
        ws.cell(r, 1, "group").font = _BOLD_FONT
        ws.cell(r, 2, "contrib (log-pt)").font = _BOLD_FONT
        ws.cell(r, 3, "contrib (%)").font = _BOLD_FONT
        ws.cell(r, 4, "pct of composition").font = _BOLD_FONT
        r += 1

        contrib_rows = []
        for g in GROUP_LABELS:
            lp = cache.get(f"{g}_comp_lp")
            if lp is not None and not np.isnan(lp):
                pct_v = (np.exp(lp) - 1.0) * 100.0
                share = pct_v / comp * 100.0 if comp else None
                contrib_rows.append((g, lp, pct_v, share))

        # Sort by absolute contribution descending
        contrib_rows.sort(key=lambda x: abs(x[2]), reverse=True)

        for g, lp_v, pct_v, share in contrib_rows:
            ws.cell(r, 1, g).font = _BODY_FONT
            ws.cell(r, 2, round(lp_v, 5)).number_format = "0.00000"
            ws.cell(r, 3, round(pct_v, 4)).number_format = _NUM3
            if share is not None:
                ws.cell(r, 4, round(share, 1)).number_format = "0.0"
            r += 1
    else:
        ws.cell(r, 1, "Group contributions: run engine with --save-cache to populate.").font = _NOTE_FONT
        r += 1

    _autosize(ws, {1: 40, 2: 18, 3: 14, 4: 20})


def _build_params_meta_sheet(
    wb: Workbook,
    rep: pd.DataFrame,
    boc: pd.Series,
    refreshed_at: str,
) -> None:
    """Sheet 4: frozen Spec, calibration stats, provenance, diagnosis."""
    ws = wb.create_sheet("params_meta")

    from pipeline.lfs_micro.spec import DEFAULT_SPEC

    # Latest available PUMF month
    rep_idx = rep.set_index("date").sort_index()
    latest_rep = rep_idx.index.max() if not rep_idx.empty else None
    pumf_vintage = latest_rep.strftime("%Y-%m") if latest_rep is not None else "unknown"

    # Compute current fit stats
    common = rep_idx.index.intersection(boc.index)
    if len(common) >= 2:
        diff = rep_idx.loc[common, "underlying_pct"] - boc.loc[common]
        rmse_full = float(np.sqrt((diff.dropna() ** 2).mean()))
        mae_full = float(diff.dropna().abs().mean())
        corr_full = float(rep_idx.loc[common, "underlying_pct"].dropna().corr(boc.loc[common]))
        n_overlap = int(diff.dropna().count())
        # Last 18 months
        cutoff = common.max() - pd.DateOffset(months=18)
        recent = diff.dropna()[diff.dropna().index >= cutoff]
        rmse_18 = float(np.sqrt((recent ** 2).mean())) if len(recent) >= 2 else float("nan")
        mae_18 = float(recent.abs().mean()) if len(recent) >= 2 else float("nan")
    else:
        rmse_full = mae_full = corr_full = rmse_18 = mae_18 = float("nan")
        n_overlap = 0

    # Diagnosis conclusion
    diagnosis = _calibration_conclusion()

    params = [
        # Spec
        ("SPEC: weighted", str(DEFAULT_SPEC.weighted)),
        ("SPEC: smoothing", DEFAULT_SPEC.smoothing),
        ("SPEC: ob_reference", DEFAULT_SPEC.ob_reference),
        ("SPEC: min_cell_count", DEFAULT_SPEC.min_cell_count),
        ("SPEC: tenure_bins", str(list(DEFAULT_SPEC.tenure_bins))),
        (None, None),
        # Calibration stats
        ("RMSE (full sample, pp)", round(rmse_full, 4)),
        ("MAE (full sample, pp)", round(mae_full, 4)),
        ("Correlation (full sample)", round(corr_full, 4)),
        ("Overlap n (months)", n_overlap),
        ("RMSE (last 18 months, pp)", round(rmse_18, 4)),
        ("MAE (last 18 months, pp)", round(mae_18, 4)),
        (None, None),
        # Provenance
        ("PUMF vintage (latest month)", pumf_vintage),
        ("BoC benchmark series", "INDINF_LFSMICRO_M (Valet)"),
        ("BoC benchmark last date", boc.index.max().strftime("%Y-%m") if not boc.empty else "unknown"),
        ("Methodology", "Oaxaca-Blinder two-fold (Bounajm/Devakos/Galassi, BoC SAN 2024-23)"),
        ("PUMF source URL", "https://www150.statcan.gc.ca/n1/pub/71m0001x/71m0001x2021001-eng.htm"),
        ("Refreshed at (UTC)", refreshed_at),
        (None, None),
        # Caveats
        ("Caveat: vintage mismatch", "BoC computed its series in real-time on pre-revision vintages. "
                                     "After Feb 2025 StatCan rebased the PUMF using NAICS 2022; both we "
                                     "and the BoC now use the revised series so this is not a source of "
                                     "current divergence. Verify: NAICS_21 codes are consistent across all "
                                     "PUMF months (confirmed in calibration)."),
        ("Caveat: log-pt conversion", "Underlying growth is computed in log-points, "
                                      "then converted via (exp(lp)-1)*100. "
                                      "For values near 3%, this differs from raw log-pts by <0.05pp."),
        ("Caveat: MA3 smoothing", "Centered 3-month MA applied to monthly O-B results "
                                  "before y/y differencing. Edge months (latest) use trailing MA or "
                                  "raw point estimate; note in headline sheet."),
        (None, None),
        # Divergence diagnosis
        ("Divergence diagnosis summary", diagnosis),
    ]

    # Column headers
    ws.cell(1, 1, "parameter").font = _HEADER_FONT
    ws.cell(1, 1).fill = _HEADER_FILL
    ws.cell(1, 2, "value").font = _HEADER_FONT
    ws.cell(1, 2).fill = _HEADER_FILL
    ws.cell(1, 1).border = _BORDER
    ws.cell(1, 2).border = _BORDER

    for ri, (key, val) in enumerate(params, start=2):
        if key is None:
            continue
        kc = ws.cell(ri, 1, key)
        kc.font = _BOLD_FONT
        kc.border = _BORDER
        vc = ws.cell(ri, 2, val)
        vc.font = _BODY_FONT
        vc.border = _BORDER
        vc.alignment = Alignment(wrap_text=True, vertical="top")

        if isinstance(val, float) and not np.isnan(val):
            vc.number_format = _NUM4 if key.startswith("RMSE") or key.startswith("MAE") else (
                _NUM4 if "Correlation" in key else None
            )

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A2"


_NUM4 = "0.0000"


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_workbook(
    out_path: str | Path = DEFAULT_OUT,
    *,
    overwrite: bool = True,
) -> Path:
    """Build the LFS-micro replication workbook.

    Args:
        out_path:  Destination .xlsx path (created/overwritten).
        overwrite: If False, raises FileExistsError when file exists.

    Returns:
        Path to the written workbook.
    """
    out_path = Path(out_path)
    if out_path.exists() and not overwrite:
        raise FileExistsError(
            f"Workbook already exists: {out_path}\n"
            "Pass overwrite=True to regenerate."
        )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    rep = _load_replication()
    boc = _load_boc()
    engine_cache = _load_engine_cache()
    refreshed_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()

    _build_headline_sheet(wb, rep, boc)
    _build_decomposition_sheet(wb, rep, engine_cache)
    _build_latest_month_sheet(wb, rep, boc, engine_cache)
    _build_params_meta_sheet(wb, rep, boc, refreshed_at)

    # Save with Windows file-lock companion fallback
    try:
        wb.save(out_path)
        wb.close()
        return out_path
    except PermissionError:
        companion = out_path.with_name(
            out_path.stem.replace("_replication", "_output") + out_path.suffix
        )
        wb.save(companion)
        wb.close()
        return companion


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None) -> int:
    import argparse
    parser = argparse.ArgumentParser(description="Build LFS-micro replication workbook")
    parser.add_argument("--out", default=str(DEFAULT_OUT), help="output xlsx path")
    args = parser.parse_args(argv)
    out = build_workbook(args.out, overwrite=True)
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
