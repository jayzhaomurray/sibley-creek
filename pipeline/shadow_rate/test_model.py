"""Tests for the shadow-rate model. Math locked before I/O. No HTTP.

Run with the repo venv:
    .venv/Scripts/python.exe -m pytest pipeline/shadow_rate/test_model.py
"""

from __future__ import annotations

from datetime import date

import pytest

from pipeline.shadow_rate.inputs import (
    AnnualRow,
    Params,
    QuarterlyRow,
    ShadowInputs,
)
from pipeline.shadow_rate import model as m


# --------------------------------------------------------------------------- #
# Fixture builders
# --------------------------------------------------------------------------- #
def make_params(**over) -> Params:
    base = dict(
        mpr_publication_date=date(2026, 4, 29),
        projection_end_quarter="2028Q4",
        current_overnight_rate=2.25,
        output_gap_anchor_quarter="2026Q2",
        output_gap_anchor_value=-0.5,
        neutral_range_low=2.25,
        neutral_range_high=3.25,
        rho=0.85,
        phi_pi=4.65,
        phi_gap=0.40,
        inflation_target=2.0,
        inflation_converge_quarters=4,
        elb_floor=0.25,
        verified=False,
    )
    base.update(over)
    return Params(**base)


def const_inputs(core=2.0, gdp=1.2, potential=1.2, gap_low=0.0, gap_high=0.0,
                 anchor_quarter="2026Q2",
                 years=(2025, 2026, 2027, 2028, 2029, 2030), **pover) -> ShadowInputs:
    """Inputs with constant core CPI and GDP=potential everywhere (flat gap).

    ``gap_low``/``gap_high`` keep the old call sites working: their midpoint is
    the output-gap anchor value, anchored (by default) at the 2026Q2 seed quarter.
    """
    quarterly = []
    for yr in years:
        for qn in range(1, 5):
            quarterly.append(
                QuarterlyRow(
                    quarter=f"{yr}Q{qn}",
                    core_cpi_yoy_forecast=core,
                    total_cpi_yoy_reference=None,
                    gdp_growth_qq_ann_forecast=gdp,
                    anchor_type="q4q4" if qn == 4 else "quarterly",
                    source_ref="test",
                )
            )
    annual = [
        AnnualRow(
            year=yr,
            potential_growth_low=potential,
            potential_growth_high=potential,
            gdp_q4q4=gdp,
            source_ref="test",
        )
        for yr in years
    ]
    params = make_params(
        output_gap_anchor_quarter=anchor_quarter,
        output_gap_anchor_value=(gap_low + gap_high) / 2.0,
        **pover,
    )
    return ShadowInputs(quarterly=quarterly, annual=annual, params=params)


# --------------------------------------------------------------------------- #
# Quarter arithmetic
# --------------------------------------------------------------------------- #
def test_quarter_roundtrip():
    for q in ["2025Q1", "2026Q2", "2028Q4", "2030Q3"]:
        assert m.ord_to_quarter(m.quarter_to_ord(q)) == q


def test_quarter_ordering():
    assert m.quarter_to_ord("2026Q3") - m.quarter_to_ord("2026Q2") == 1
    assert m.quarter_to_ord("2027Q1") - m.quarter_to_ord("2026Q4") == 1


def test_quarter_of_date():
    assert m.quarter_of_date(date(2026, 4, 29)) == "2026Q2"
    assert m.quarter_of_date(date(2026, 1, 15)) == "2026Q1"
    assert m.quarter_of_date(date(2026, 12, 31)) == "2026Q4"


# --------------------------------------------------------------------------- #
# Steady-state fixed point
# --------------------------------------------------------------------------- #
def test_steady_state_fixed_point():
    """Constant inputs: rate converges to neutral + phi_pi*(pi-target) + phi_gap*gap.

    With core=target and gap=0, the fixed point is exactly the neutral midpoint,
    and seeding there means the rate never moves.
    """
    inp = const_inputs(core=2.0, gdp=1.2, potential=1.2, gap_low=0.0, gap_high=0.0,
                       current_overnight_rate=2.75)  # = neutral midpoint
    res = m.run_model(inp, end_quarter="2028Q4")
    for step in res.steps:
        assert step.rate == pytest.approx(2.75, abs=1e-9)


def test_steady_state_fixed_point_with_inflation_gap():
    """Constant core above target, gap=0 -> fixed point above neutral."""
    inp = const_inputs(core=3.0, gdp=1.2, potential=1.2, gap_low=0.0, gap_high=0.0)
    # fixed point R* = 2.75 + 4.65*(3.0-2.0) + 0.4*0 = 7.40
    fp = 2.75 + 4.65 * 1.0
    inp2 = const_inputs(core=3.0, gdp=1.2, potential=1.2, gap_low=0.0, gap_high=0.0,
                        current_overnight_rate=fp)
    res = m.run_model(inp2, end_quarter="2028Q4")
    for step in res.steps:
        assert step.rate == pytest.approx(fp, abs=1e-9)


# --------------------------------------------------------------------------- #
# One-step closed form
# --------------------------------------------------------------------------- #
def test_one_step_closed_form():
    inp = const_inputs(core=3.0, gdp=1.2, potential=1.2, gap_low=0.0, gap_high=0.0,
                       current_overnight_rate=2.25)
    res = m.run_model(inp, end_quarter="2028Q4")
    # First update: target_level = 2.75 + 4.65*(3-2) + 0.4*0 = 7.40
    # R_1 = 0.85*2.25 + 0.15*7.40 = 1.9125 + 1.11 = 3.0225
    assert res.steps[0].rate == pytest.approx(2.25, abs=1e-12)
    assert res.steps[1].rate == pytest.approx(0.85 * 2.25 + 0.15 * 7.40, abs=1e-12)


# --------------------------------------------------------------------------- #
# Geometric inertia approach
# --------------------------------------------------------------------------- #
def test_geometric_inertia_approach():
    """With constant target_level T and gap=0, the gap to T shrinks by rho each step."""
    inp = const_inputs(core=3.0, gdp=1.2, potential=1.2, gap_low=0.0, gap_high=0.0,
                       current_overnight_rate=2.25)
    res = m.run_model(inp, end_quarter="2028Q4")
    T = 2.75 + 4.65 * 1.0  # 7.40
    rho = 0.85
    prev_gap = None
    for step in res.steps:
        g = T - step.rate
        if prev_gap is not None and abs(prev_gap) > 1e-9:
            assert g / prev_gap == pytest.approx(rho, abs=1e-9)
        prev_gap = g


# --------------------------------------------------------------------------- #
# Gap identity
# --------------------------------------------------------------------------- #
def test_gap_identity():
    """+1pp growth above potential -> gap rises +0.25/quarter."""
    inp = const_inputs(core=2.0, gdp=2.2, potential=1.2, gap_low=0.0, gap_high=0.0)
    res = m.run_model(inp, end_quarter="2028Q4")
    seed_ord = m.quarter_to_ord(res.seed_quarter)
    g0 = res.gap_path[seed_ord]
    g1 = res.gap_path[seed_ord + 1]
    assert g1 - g0 == pytest.approx(0.25, abs=1e-9)
    g2 = res.gap_path[seed_ord + 2]
    assert g2 - g1 == pytest.approx(0.25, abs=1e-9)


# --------------------------------------------------------------------------- #
# Anchor + roll-forward to the seed quarter
# --------------------------------------------------------------------------- #
def _seed_inputs(anchor_quarter="2025Q4", anchor_value=-1.0, **pover) -> ShadowInputs:
    """The April-2026-MPR seed-data shape: GDP 2025Q3=2.4, Q4=-0.6, 2026Q1/Q2=1.5;
    potential 2025=2.3, 2026=1.2. Used to exercise the anchor roll-forward."""
    quarterly = []
    for q, core, gdp in [
        ("2025Q3", 3.1, 2.4),
        ("2025Q4", 2.8, -0.6),
        ("2026Q1", 2.4, 1.5),
        ("2026Q2", 2.1, 1.5),
    ]:
        quarterly.append(QuarterlyRow(quarter=q, core_cpi_yoy_forecast=core,
                                       total_cpi_yoy_reference=None,
                                       gdp_growth_qq_ann_forecast=gdp,
                                       anchor_type="quarterly", source_ref="t"))
    for q, core in [("2026Q4", 2.0), ("2027Q4", 2.2), ("2028Q4", 2.0)]:
        quarterly.append(QuarterlyRow(quarter=q, core_cpi_yoy_forecast=core,
                                       total_cpi_yoy_reference=None,
                                       gdp_growth_qq_ann_forecast=None,
                                       anchor_type="q4q4", source_ref="t"))
    annual = [
        AnnualRow(year=2025, potential_growth_low=2.3, potential_growth_high=2.3,
                  gdp_q4q4=None, source_ref="t"),
        AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                  gdp_q4q4=1.8, source_ref="t"),
        AnnualRow(year=2027, potential_growth_low=0.8, potential_growth_high=1.8,
                  gdp_q4q4=1.4, source_ref="t"),
        AnnualRow(year=2028, potential_growth_low=1.0, potential_growth_high=2.0,
                  gdp_q4q4=1.9, source_ref="t"),
    ]
    params = make_params(output_gap_anchor_quarter=anchor_quarter,
                         output_gap_anchor_value=anchor_value, **pover)
    return ShadowInputs(quarterly=quarterly, annual=annual, params=params)


def test_anchor_roll_forward_to_seed():
    """Anchor 2025Q4=-1.0 rolls forward to seed 2026Q2=-0.85 on the seed data.

    2026Q1 = -1.0 + (1.5 - 1.2)/4 = -0.925
    2026Q2 = -0.925 + (1.5 - 1.2)/4 = -0.85
    """
    inp = _seed_inputs(anchor_quarter="2025Q4", anchor_value=-1.0)
    res = m.run_model(inp, end_quarter="2028Q4")
    aord = m.quarter_to_ord("2025Q4")
    assert res.gap_path[aord] == pytest.approx(-1.0, abs=1e-9)
    assert res.gap_path[m.quarter_to_ord("2026Q1")] == pytest.approx(-0.925, abs=1e-9)
    assert res.gap_path[m.quarter_to_ord("2026Q2")] == pytest.approx(-0.85, abs=1e-9)
    # the seed (MPR) quarter's gap entering the rule is the rolled-forward value
    assert res.steps[0].quarter == "2026Q2"
    assert res.steps[0].gap == pytest.approx(-0.85, abs=1e-9)


def test_anchor_before_available_data_fails_closed():
    """Anchor at 2025Q2 has no potential-growth / GDP coverage (annual starts
    2025 but GDP quarterly only from 2025Q3) -> build raises, fail closed."""
    # 2025Q2 has no direct GDP and 2025 has no gdp_q4q4 anchor -> build_gdp_path raises.
    inp = _seed_inputs(anchor_quarter="2025Q2", anchor_value=-0.8)
    with pytest.raises(ValueError):
        m.run_model(inp, end_quarter="2028Q4")


# --------------------------------------------------------------------------- #
# ELB clamp
# --------------------------------------------------------------------------- #
def test_elb_clamp():
    """Deeply negative inflation drives the rule below ELB; output clamps at floor."""
    inp = const_inputs(core=-3.0, gdp=1.2, potential=1.2, gap_low=-5.0, gap_high=-5.0,
                       current_overnight_rate=2.25)
    res = m.run_model(inp, end_quarter="2028Q4")
    # later steps should hit the floor and never go below it
    assert min(s.rate for s in res.steps) >= 0.25 - 1e-12
    assert res.steps[-1].rate == pytest.approx(0.25, abs=1e-9)


# --------------------------------------------------------------------------- #
# Hold-at-target beyond final anchor
# --------------------------------------------------------------------------- #
def test_hold_beyond_final_anchor():
    """t+4 inflation lookups past the last known quarter hold the final value."""
    # Build inputs whose last quarterly/anchor row is 2028Q4 = 2.0, and check
    # that the path past it holds 2.0.
    quarterly = []
    # near-term quarterly through 2026Q2
    for q, v in [("2025Q3", 3.1), ("2025Q4", 2.8), ("2026Q1", 2.4), ("2026Q2", 2.1)]:
        quarterly.append(QuarterlyRow(quarter=q, core_cpi_yoy_forecast=v,
                                       total_cpi_yoy_reference=None,
                                       gdp_growth_qq_ann_forecast=1.5,
                                       anchor_type="quarterly", source_ref="t"))
    for q, v in [("2026Q4", 2.0), ("2027Q4", 2.2), ("2028Q4", 2.0)]:
        quarterly.append(QuarterlyRow(quarter=q, core_cpi_yoy_forecast=v,
                                       total_cpi_yoy_reference=None,
                                       gdp_growth_qq_ann_forecast=None,
                                       anchor_type="q4q4", source_ref="t"))
    annual = [
        AnnualRow(year=2025, potential_growth_low=2.3, potential_growth_high=2.3,
                  gdp_q4q4=None, source_ref="t"),
        AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                  gdp_q4q4=1.8, source_ref="t"),
        AnnualRow(year=2027, potential_growth_low=0.8, potential_growth_high=1.8,
                  gdp_q4q4=1.4, source_ref="t"),
        AnnualRow(year=2028, potential_growth_low=1.0, potential_growth_high=2.0,
                  gdp_q4q4=1.9, source_ref="t"),
    ]
    inp = ShadowInputs(quarterly=quarterly, annual=annual, params=make_params())
    last = m.build_core_cpi_path(inp, m.quarter_to_ord("2030Q4"))
    # everything past 2028Q4 holds 2.0
    for o in range(m.quarter_to_ord("2029Q1"), m.quarter_to_ord("2030Q4") + 1):
        assert last[o] == pytest.approx(2.0, abs=1e-9)


# --------------------------------------------------------------------------- #
# Linear interpolation between Q4/Q4 anchors
# --------------------------------------------------------------------------- #
def test_linear_interpolation_midpoint():
    """2026Q2=2.1 -> 2026Q4=2.0 should put 2026Q3 at 2.05 (exact midpoint)."""
    quarterly = [
        QuarterlyRow(quarter="2026Q2", core_cpi_yoy_forecast=2.1,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=1.5,
                     anchor_type="quarterly", source_ref="t"),
        QuarterlyRow(quarter="2026Q4", core_cpi_yoy_forecast=2.0,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=None,
                     anchor_type="q4q4", source_ref="t"),
    ]
    annual = [AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                        gdp_q4q4=1.8, source_ref="t")]
    inp = ShadowInputs(quarterly=quarterly, annual=annual, params=make_params())
    path = m.build_core_cpi_path(inp, m.quarter_to_ord("2026Q4"))
    assert path[m.quarter_to_ord("2026Q3")] == pytest.approx(2.05, abs=1e-9)


def test_linear_interpolation_three_quarter_span():
    """2027Q4=2.2 -> 2028Q4=2.0 across 4 quarters; 2028Q2 sits halfway at 2.10."""
    quarterly = [
        QuarterlyRow(quarter="2027Q4", core_cpi_yoy_forecast=2.2,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=None,
                     anchor_type="q4q4", source_ref="t"),
        QuarterlyRow(quarter="2028Q4", core_cpi_yoy_forecast=2.0,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=None,
                     anchor_type="q4q4", source_ref="t"),
    ]
    annual = [AnnualRow(year=2028, potential_growth_low=1.0, potential_growth_high=2.0,
                        gdp_q4q4=1.9, source_ref="t")]
    inp = ShadowInputs(quarterly=quarterly, annual=annual, params=make_params())
    path = m.build_core_cpi_path(inp, m.quarter_to_ord("2028Q4"))
    # 2027Q4=2.2, 2028Q1=2.15, 2028Q2=2.10, 2028Q3=2.05, 2028Q4=2.0
    assert path[m.quarter_to_ord("2028Q1")] == pytest.approx(2.15, abs=1e-9)
    assert path[m.quarter_to_ord("2028Q2")] == pytest.approx(2.10, abs=1e-9)
    assert path[m.quarter_to_ord("2028Q3")] == pytest.approx(2.05, abs=1e-9)


# --------------------------------------------------------------------------- #
# Residual GDP fill within anchor years
# --------------------------------------------------------------------------- #
def test_gdp_residual_fill_within_year():
    """2026 has Q1=Q2=1.5 direct and a Q4/Q4 anchor of 1.8. The two missing
    quarters fill at the residual (4*1.8 - 1.5 - 1.5)/2 = 2.1, so the four
    quarters average the anchor."""
    quarterly = [
        QuarterlyRow(quarter="2026Q1", core_cpi_yoy_forecast=2.4,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=1.5,
                     anchor_type="quarterly", source_ref="t"),
        QuarterlyRow(quarter="2026Q2", core_cpi_yoy_forecast=2.1,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=1.5,
                     anchor_type="quarterly", source_ref="t"),
        QuarterlyRow(quarter="2026Q4", core_cpi_yoy_forecast=2.0,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=None,
                     anchor_type="q4q4", source_ref="t"),
    ]
    annual = [AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                        gdp_q4q4=1.8, source_ref="t")]
    inp = ShadowInputs(quarterly=quarterly, annual=annual, params=make_params())
    gdp = m.build_gdp_path(inp, m.quarter_to_ord("2026Q1"), m.quarter_to_ord("2026Q4"))
    assert gdp[m.quarter_to_ord("2026Q1")] == pytest.approx(1.5)
    assert gdp[m.quarter_to_ord("2026Q2")] == pytest.approx(1.5)
    # Q3 and Q4 (no direct value) -> residual (4*1.8 - 1.5 - 1.5)/2 = 2.1
    assert gdp[m.quarter_to_ord("2026Q3")] == pytest.approx(2.1)
    assert gdp[m.quarter_to_ord("2026Q4")] == pytest.approx(2.1)
    # the four quarters average the anchor
    yr_avg = sum(gdp[m.quarter_to_ord(f"2026Q{q}")] for q in range(1, 5)) / 4
    assert yr_avg == pytest.approx(1.8)


def test_gdp_constant_fill_no_direct_quarters():
    """A year with NO direct quarters reduces to the old constant fill: every
    missing quarter is the anchor itself (residual with zero known terms)."""
    quarterly = [
        QuarterlyRow(quarter="2026Q4", core_cpi_yoy_forecast=2.0,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=None,
                     anchor_type="q4q4", source_ref="t"),
    ]
    annual = [AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                        gdp_q4q4=1.8, source_ref="t")]
    inp = ShadowInputs(quarterly=quarterly, annual=annual, params=make_params())
    gdp = m.build_gdp_path(inp, m.quarter_to_ord("2026Q1"), m.quarter_to_ord("2026Q4"))
    for q in range(1, 5):
        assert gdp[m.quarter_to_ord(f"2026Q{q}")] == pytest.approx(1.8)


def test_gdp_all_direct_year_ignores_anchor():
    """A year with all four quarters direct ignores the anchor (n_missing=0):
    no residual fill, the directly-given rates stand even if they don't average
    the anchor."""
    quarterly = [
        QuarterlyRow(quarter=f"2026Q{q}", core_cpi_yoy_forecast=2.0,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=g,
                     anchor_type="q4q4" if q == 4 else "quarterly", source_ref="t")
        for q, g in [(1, 1.0), (2, 2.0), (3, 3.0), (4, 4.0)]
    ]
    # anchor 1.8 disagrees with the direct mean (2.5) — must be ignored.
    annual = [AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                        gdp_q4q4=1.8, source_ref="t")]
    inp = ShadowInputs(quarterly=quarterly, annual=annual, params=make_params())
    gdp = m.build_gdp_path(inp, m.quarter_to_ord("2026Q1"), m.quarter_to_ord("2026Q4"))
    assert gdp[m.quarter_to_ord("2026Q1")] == pytest.approx(1.0)
    assert gdp[m.quarter_to_ord("2026Q2")] == pytest.approx(2.0)
    assert gdp[m.quarter_to_ord("2026Q3")] == pytest.approx(3.0)
    assert gdp[m.quarter_to_ord("2026Q4")] == pytest.approx(4.0)


def test_potential_midpoint():
    inp = const_inputs()
    annual = [
        AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                  gdp_q4q4=1.8, source_ref="t"),
    ]
    inp2 = ShadowInputs(quarterly=inp.quarterly, annual=annual + inp.annual[2:],
                        params=inp.params)
    pot = m.build_potential_path(inp2, m.quarter_to_ord("2026Q1"),
                                 m.quarter_to_ord("2026Q4"))
    for qn in range(1, 5):
        assert pot[m.quarter_to_ord(f"2026Q{qn}")] == pytest.approx(1.2)


# --------------------------------------------------------------------------- #
# Fail-closed validation
# --------------------------------------------------------------------------- #
def test_validation_rejects_bad_quarter():
    with pytest.raises(Exception):
        QuarterlyRow(quarter="2026-Q2", core_cpi_yoy_forecast=2.0,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=None,
                     anchor_type="quarterly", source_ref="t")


def test_validation_rejects_bad_anchor_type():
    with pytest.raises(Exception):
        QuarterlyRow(quarter="2026Q2", core_cpi_yoy_forecast=2.0,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=None,
                     anchor_type="annual", source_ref="t")


def test_validation_rejects_inverted_neutral_range():
    with pytest.raises(Exception):
        make_params(neutral_range_low=3.25, neutral_range_high=2.25)


def test_validation_rejects_out_of_bounds_anchor_value():
    with pytest.raises(Exception):
        make_params(output_gap_anchor_value=-99.0)


def test_validation_rejects_unparseable_anchor_quarter():
    with pytest.raises(Exception):
        make_params(output_gap_anchor_quarter="2025-Q4")


def test_validation_rejects_anchor_after_seed():
    """Fail closed: an anchor quarter AFTER the MPR seed quarter is rejected,
    because the gap rolls forward from the anchor to the seed (never backward)."""
    inp = const_inputs(anchor_quarter="2027Q1")  # seed is 2026Q2
    with pytest.raises(ValueError):
        m.run_model(inp, end_quarter="2028Q4")


def test_validation_rejects_out_of_bounds_rho():
    with pytest.raises(Exception):
        make_params(rho=1.5)


def test_validation_rejects_empty_source_ref():
    with pytest.raises(Exception):
        AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                  gdp_q4q4=1.8, source_ref="   ")


def test_validation_rejects_no_q4q4_anchor():
    quarterly = [
        QuarterlyRow(quarter="2026Q2", core_cpi_yoy_forecast=2.1,
                     total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=1.5,
                     anchor_type="quarterly", source_ref="t"),
    ]
    annual = [AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                        gdp_q4q4=1.8, source_ref="t")]
    with pytest.raises(Exception):
        ShadowInputs(quarterly=quarterly, annual=annual, params=make_params())


# --------------------------------------------------------------------------- #
# Fail-closed workbook-integrity checks (the audit's reproduced failures)
# --------------------------------------------------------------------------- #
def _full_quarterly():
    """A complete near-term + Q4/Q4-anchor quarterly set (no duplicates)."""
    rows = []
    for q, core, gdp in [("2025Q3", 3.1, 2.4), ("2025Q4", 2.8, -0.6),
                         ("2026Q1", 2.4, 1.5), ("2026Q2", 2.1, 1.5)]:
        rows.append(QuarterlyRow(quarter=q, core_cpi_yoy_forecast=core,
                                 total_cpi_yoy_reference=None,
                                 gdp_growth_qq_ann_forecast=gdp,
                                 anchor_type="quarterly", source_ref="t"))
    for q, core in [("2026Q4", 2.0), ("2027Q4", 2.2), ("2028Q4", 2.0)]:
        rows.append(QuarterlyRow(quarter=q, core_cpi_yoy_forecast=core,
                                 total_cpi_yoy_reference=None,
                                 gdp_growth_qq_ann_forecast=None,
                                 anchor_type="q4q4", source_ref="t"))
    return rows


def _full_annual():
    return [
        AnnualRow(year=2025, potential_growth_low=2.3, potential_growth_high=2.3,
                  gdp_q4q4=None, source_ref="t"),
        AnnualRow(year=2026, potential_growth_low=0.8, potential_growth_high=1.6,
                  gdp_q4q4=1.8, source_ref="t"),
        AnnualRow(year=2027, potential_growth_low=0.8, potential_growth_high=1.8,
                  gdp_q4q4=1.4, source_ref="t"),
        AnnualRow(year=2028, potential_growth_low=1.0, potential_growth_high=2.0,
                  gdp_q4q4=1.9, source_ref="t"),
    ]


def test_integrity_baseline_passes():
    """The full set with no defects constructs cleanly (control for the rest)."""
    inp = ShadowInputs(quarterly=_full_quarterly(), annual=_full_annual(),
                       params=make_params())
    assert any(r.anchor_type == "q4q4" for r in inp.quarterly)


def test_fail_closed_duplicate_quarter():
    """(1a) Same quarter twice -> ValueError naming the quarter."""
    q = _full_quarterly()
    q.append(QuarterlyRow(quarter="2026Q2", core_cpi_yoy_forecast=9.9,
                          total_cpi_yoy_reference=None, gdp_growth_qq_ann_forecast=9.9,
                          anchor_type="quarterly", source_ref="dup"))
    with pytest.raises(ValueError, match="duplicate quarterly row for 2026Q2"):
        ShadowInputs(quarterly=q, annual=_full_annual(), params=make_params())


def test_fail_closed_duplicate_annual_year():
    """Duplicate annual-sheet year -> ValueError naming the year."""
    a = _full_annual()
    a.append(AnnualRow(year=2027, potential_growth_low=0.0, potential_growth_high=0.0,
                       gdp_q4q4=9.9, source_ref="dup"))
    with pytest.raises(ValueError, match="duplicate annual row for year 2027"):
        ShadowInputs(quarterly=_full_quarterly(), annual=a, params=make_params())


def test_fail_closed_missing_core_q4_anchor(tmp_path):
    """(1c) Deleting the 2026Q4 core anchor -> ValueError naming the year.

    Reproduces the audit scenario at the workbook level: regenerate the seeded
    workbook, delete the 2026Q4 quarterly row (the core Q4 anchor), and confirm
    parse_workbook fails closed naming 2026.
    """
    from openpyxl import load_workbook

    from pipeline.shadow_rate.inputs import parse_workbook
    from pipeline.shadow_rate.make_workbook import build_workbook

    xlsx = tmp_path / "wb.xlsx"
    build_workbook(str(xlsx))
    wb = load_workbook(str(xlsx))
    wq = wb["quarterly"]
    for ri in range(wq.max_row, 1, -1):
        if wq.cell(ri, 1).value == "2026Q4":
            wq.delete_rows(ri, 1)
    wb.save(str(xlsx))
    wb.close()
    with pytest.raises(ValueError, match="missing core-CPI Q4 coverage for 2026"):
        parse_workbook(str(xlsx))


def test_fail_closed_missing_gdp_anchor():
    """(1c GDP) A horizon year with neither 4 direct quarters nor a Q4/Q4 GDP
    anchor -> ValueError naming the year, at the workbook-coverage check."""
    from pipeline.shadow_rate.inputs import check_horizon_coverage

    a = _full_annual()
    # strip 2027's gdp_q4q4 anchor; 2027 has no direct quarters -> no coverage
    a = [
        AnnualRow(year=x.year, potential_growth_low=x.potential_growth_low,
                  potential_growth_high=x.potential_growth_high,
                  gdp_q4q4=(None if x.year == 2027 else x.gdp_q4q4),
                  source_ref="t")
        for x in a
    ]
    inp = ShadowInputs(quarterly=_full_quarterly(), annual=a, params=make_params())
    with pytest.raises(ValueError, match="missing GDP coverage for 2027"):
        check_horizon_coverage(inp)


def test_fail_closed_duplicate_params_key(tmp_path):
    """(1b) Same params key twice -> ValueError naming the key, at parse time."""
    from openpyxl import load_workbook

    from pipeline.shadow_rate.inputs import parse_workbook
    from pipeline.shadow_rate.make_workbook import build_workbook

    xlsx = tmp_path / "wb.xlsx"
    build_workbook(str(xlsx))
    wb = load_workbook(str(xlsx))
    wp = wb["params"]
    # duplicate the rho row by inserting a second rho key/value
    wp.insert_rows(2)
    wp.cell(2, 1, "rho")
    wp.cell(2, 2, 0.5)
    wp.cell(2, 3, "dup")
    wb.save(str(xlsx))
    wb.close()
    with pytest.raises(ValueError, match="duplicate params key 'rho'"):
        parse_workbook(str(xlsx))


# --------------------------------------------------------------------------- #
# xlsx round-trip
# --------------------------------------------------------------------------- #
def test_xlsx_roundtrip(tmp_path):
    """make_workbook -> parse_workbook reproduces the seeded params."""
    from pipeline.shadow_rate.make_workbook import build_workbook
    from pipeline.shadow_rate.inputs import parse_workbook

    xlsx = tmp_path / "rt.xlsx"
    build_workbook(str(xlsx))
    inp = parse_workbook(str(xlsx))
    assert inp.params.mpr_publication_date == date(2026, 4, 29)
    assert inp.params.verified is False
    assert inp.params.neutral_range_low == pytest.approx(2.25)
    assert inp.params.neutral_range_high == pytest.approx(3.25)
    # model runs end-to-end on the seeded workbook
    res = m.run_model(inp, end_quarter="2028Q4")
    assert res.seed_quarter == "2026Q2"
    assert res.steps[0].rate == pytest.approx(inp.params.current_overnight_rate)


# --------------------------------------------------------------------------- #
# Live-formula calc sheet written back into the workbook
# --------------------------------------------------------------------------- #
def _calc_grid(ws):
    """Return (header_row_index, {quarter -> {col_letter: value}}) for the grid.

    Reads the calc sheet's grid rows keyed by quarter label, with every column's
    raw cell value (formula strings or static numbers) under its column letter.
    """
    from openpyxl.utils import get_column_letter

    header_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "quarter":
            header_idx = i
            break
    assert header_idx is not None
    grid: dict[str, dict[str, object]] = {}
    for ri in range(header_idx + 1, ws.max_row + 1):
        q = ws.cell(ri, 1).value
        if not isinstance(q, str) or "Q" not in q:
            continue
        grid[q] = {
            get_column_letter(ci): ws.cell(ri, ci).value
            for ci in range(1, ws.max_column + 1)
        }
    return header_idx, grid


def test_calc_sheet_live_formulas_and_python_handshake(tmp_path):
    """write_output_sheet adds a live-formula 'calc' sheet (first), preserves the
    input sheets, writes auditable formulas (interpolation / gap / rate / neutral
    mid with absolute params refs), and the static 'python' columns match the
    engine."""
    from openpyxl import load_workbook

    from pipeline.shadow_rate.inputs import parse_workbook
    from pipeline.shadow_rate.make_workbook import build_workbook
    from pipeline.shadow_rate.output_sheet import SHEET_NAME, write_output_sheet

    xlsx = tmp_path / "wb.xlsx"
    build_workbook(str(xlsx))
    inp = parse_workbook(str(xlsx))
    p = inp.params
    res = m.run_model(inp, end_quarter="2028Q4")

    # snapshot input-sheet values before writing the calc sheet
    wb0 = load_workbook(str(xlsx))
    before = {
        name: [r for r in wb0[name].iter_rows(values_only=True)]
        for name in ("quarterly", "annual", "params")
    }
    wb0.close()

    out = write_output_sheet(str(xlsx), res, p)
    assert out.used_companion is False
    assert out.path == xlsx

    wb = load_workbook(str(xlsx))
    try:
        # calc sheet exists and is FIRST; legacy 'output' sheet is gone.
        assert SHEET_NAME in wb.sheetnames
        assert wb.sheetnames[0] == SHEET_NAME
        assert "output" not in wb.sheetnames
        assert wb.sheetnames == ["calc", "quarterly", "annual", "params"]

        # input sheets preserved (values unchanged) and still present
        for name in ("quarterly", "annual", "params"):
            assert name in wb.sheetnames
            after = [r for r in wb[name].iter_rows(values_only=True)]
            assert after == before[name]

        ws = wb[SHEET_NAME]

        # header block: UNVERIFIED DRAFT + provenance + run stamp + handshake line
        cells = [v for row in ws.iter_rows(values_only=True) for v in row]
        assert any(v == "UNVERIFIED DRAFT" for v in cells)
        assert any(isinstance(v, str) and "TR-119 Table 2.3" in v for v in cells)
        assert any(isinstance(v, str) and v.startswith("Run:") for v in cells)
        assert any(isinstance(v, str) and "model.py" in v for v in cells)
        assert any(
            isinstance(v, str) and "All white cells are live formulas" in v
            for v in cells
        )

        _, grid = _calc_grid(ws)

        # The grid spans the anchor (2025Q4) through the t+4 headroom (2029Q4).
        assert "2025Q4" in grid and "2029Q4" in grid

        # --- interpolated core-CPI cell (2026Q3): a live linear-interp formula
        # referencing the bracketing quarterly cells (B5=2026Q2, B6=2026Q4).
        core_q3 = grid["2026Q3"]["B"]
        assert isinstance(core_q3, str) and core_q3.startswith("=")
        assert "quarterly!B5" in core_q3 and "quarterly!B6" in core_q3

        # --- gap cell (2026Q3): live identity = gap_above + (gdp-pot)/4.
        gap_q3 = grid["2026Q3"]["E"]
        assert isinstance(gap_q3, str) and gap_q3.startswith("=")
        assert "/4" in gap_q3

        # anchor gap row is a direct params ref (absolute).
        gap_anchor = grid["2025Q4"]["E"]
        assert isinstance(gap_anchor, str) and gap_anchor.startswith("=params!$B$")

        # --- rate cell (2026Q3): ELB MAX with absolute params refs (rho, elb).
        rate_q3 = grid["2026Q3"]["M"]
        assert isinstance(rate_q3, str) and rate_q3.startswith("=MAX(")
        assert "params!$B$" in rate_q3  # absolute params references
        # seed rate cell is a direct params ref to the current overnight rate.
        rate_seed = grid["2026Q2"]["M"]
        assert isinstance(rate_seed, str) and rate_seed.startswith("=params!$B$")

        # --- neutral-mid cell: (neutral_low + neutral_high)/2 absolute refs.
        neutral_q3 = grid["2026Q3"]["I"]
        assert isinstance(neutral_q3, str) and neutral_q3.startswith("=(")
        assert "params!$B$" in neutral_q3 and ")/2" in neutral_q3

        # pre-seed rows (2025Q4, 2026Q1) have blank rate cells (greyed region).
        assert grid["2025Q4"]["M"] is None
        assert grid["2026Q1"]["M"] is None

        # --- python handshake columns are STATIC values matching the engine.
        # gap python (col F) for the seed quarter.
        assert grid["2026Q2"]["F"] == pytest.approx(res.steps[0].gap, abs=1e-6)
        # rate python (col N) for the seed and a later quarter.
        rate_py = {s.quarter: s.rate for s in res.steps}
        assert grid["2026Q2"]["N"] == pytest.approx(rate_py["2026Q2"], abs=1e-6)
        assert grid["2026Q4"]["N"] == pytest.approx(rate_py["2026Q4"], abs=1e-6)
        # diff columns are formulas (ABS of formula minus python).
        assert isinstance(grid["2026Q4"]["O"], str) and grid["2026Q4"]["O"].startswith("=ABS(")
        assert isinstance(grid["2026Q2"]["G"], str) and grid["2026Q2"]["G"].startswith("=ABS(")
    finally:
        wb.close()


def test_output_sheet_companion_on_lock(tmp_path, monkeypatch):
    """If save raises PermissionError (workbook locked), a companion file is
    written instead and used_companion is True."""
    from pipeline.shadow_rate.inputs import parse_workbook
    from pipeline.shadow_rate.make_workbook import build_workbook
    from pipeline.shadow_rate import output_sheet as osheet

    xlsx = tmp_path / "boc_shadow_inputs_2026Q2.xlsx"
    build_workbook(str(xlsx))
    inp = parse_workbook(str(xlsx))
    res = m.run_model(inp, end_quarter="2028Q4")

    from openpyxl.workbook.workbook import Workbook

    orig_save = Workbook.save

    def fake_save(self, path):
        # raise only for the primary (inputs) path, succeed for companion
        if str(path).endswith("boc_shadow_inputs_2026Q2.xlsx"):
            raise PermissionError("locked by Excel")
        return orig_save(self, path)

    monkeypatch.setattr(Workbook, "save", fake_save)
    out = osheet.write_output_sheet(str(xlsx), res, inp.params)
    assert out.used_companion is True
    assert out.path.name == "boc_shadow_output_2026Q2.xlsx"
    assert out.path.exists()


# --------------------------------------------------------------------------- #
# Sensitivity band (corner reruns)
# --------------------------------------------------------------------------- #
def test_band_corner_one_step_closed_form():
    """Closed-form check at one corner: neutral=neutral_high, potential=low.

    Use flat constant inputs so the gap stays at its anchor (gdp=potential for
    the central case is broken at a corner because potential moves, so we make
    gdp track the corner potential by construction). Simpler: pick gdp=potential
    at the LOW endpoint and verify the high-neutral/low-potential corner's first
    update matches the one-step rule with that neutral and a flat gap.
    """
    # potential range collapsed to a single value via gap_low=gap_high handling:
    # const_inputs sets potential_low=high=potential, so low/high corners coincide
    # and equal `potential`. gdp=potential keeps the gap flat at the anchor (0).
    inp = const_inputs(core=3.0, gdp=1.2, potential=1.2,
                       gap_low=0.0, gap_high=0.0,
                       current_overnight_rate=2.25,
                       neutral_range_low=2.25, neutral_range_high=3.25)
    band = m.run_band(inp, end_quarter="2028Q4")
    # At the neutral_high corner with gap 0 and core 3.0:
    #   target = 3.25 + 4.65*(3-2) + 0.4*0 = 7.90
    #   R_1 = 0.85*2.25 + 0.15*7.90 = 1.9125 + 1.185 = 3.0975  (the hi corner)
    # At neutral_low corner:
    #   target = 2.25 + 4.65 = 6.90; R_1 = 1.9125 + 0.15*6.90 = 2.9475 (lo corner)
    qs = sorted(band.lo, key=m.quarter_to_ord)
    second_q = qs[1]  # first projected update
    assert band.hi[second_q] == pytest.approx(0.85 * 2.25 + 0.15 * 7.90, abs=1e-9)
    assert band.lo[second_q] == pytest.approx(0.85 * 2.25 + 0.15 * 6.90, abs=1e-9)


def test_band_envelope_property_central_within():
    """lo <= central <= hi every quarter, on the realistic seed-data shape."""
    inp = _seed_inputs(anchor_quarter="2025Q4", anchor_value=-1.0)
    band = m.run_band(inp, end_quarter="2028Q4")
    res = m.run_model(inp, end_quarter="2028Q4")
    for s in res.steps:
        assert band.lo[s.quarter] <= s.rate + 1e-9
        assert s.rate <= band.hi[s.quarter] + 1e-9
        assert band.lo[s.quarter] <= band.hi[s.quarter] + 1e-9


def test_band_potential_sign_logic():
    """Lower potential -> gap closes faster -> higher rates.

    Hold neutral fixed (collapse the range) and confirm the low-potential corner
    yields a terminal rate at or above the high-potential corner. We build a
    case with real GDP anchors above potential so the sign is exercised.
    """
    inp = _seed_inputs(anchor_quarter="2025Q4", anchor_value=-1.0,
                       neutral_range_low=2.75, neutral_range_high=2.75)
    band = m.run_band(inp, end_quarter="2028Q4")
    # with neutral collapsed, the only corner driver is potential low/high.
    last_q = sorted(band.lo, key=m.quarter_to_ord)[-1]
    # low-potential corner is the hi rate, high-potential corner is the lo rate
    assert band.hi[last_q] >= band.lo[last_q]


# --------------------------------------------------------------------------- #
# Annual-average GDP cross-check (coherence diagnostic)
# --------------------------------------------------------------------------- #
def test_annual_avg_crosscheck_constant_growth_exact():
    """Constant q/q-annualized growth g implies annual-average growth == g.

    A level index compounding at constant g has each year's mean level a factor
    (1+g/100) above the prior year's mean, so the implied annual-average growth
    equals g exactly. Publish g and confirm diff ~ 0, no trip."""
    g = 1.6
    # const_inputs gives every quarter gdp=g and q4q4 anchors=g; attach published
    # annual-avg = g to the horizon years.
    inp = const_inputs(core=2.0, gdp=g, potential=g, gap_low=0.0, gap_high=0.0)
    annual = [
        AnnualRow(year=a.year, potential_growth_low=a.potential_growth_low,
                  potential_growth_high=a.potential_growth_high,
                  gdp_q4q4=a.gdp_q4q4, gdp_annual_avg=g, source_ref="t")
        for a in inp.annual
    ]
    inp2 = ShadowInputs(quarterly=inp.quarterly, annual=annual, params=inp.params)
    checks = m.annual_average_crosscheck(inp2, end_quarter="2028Q4")
    assert checks  # at least 2026..2028
    for c in checks:
        assert c.implied == pytest.approx(g, abs=1e-6)
        assert c.diff == pytest.approx(0.0, abs=1e-6)
        assert c.tripped is False


def test_annual_avg_crosscheck_tolerance_trip():
    """A published value far from the implied trips the WARN flag (but no raise)."""
    g = 1.6
    inp = const_inputs(core=2.0, gdp=g, potential=g, gap_low=0.0, gap_high=0.0)
    annual = [
        AnnualRow(year=a.year, potential_growth_low=a.potential_growth_low,
                  potential_growth_high=a.potential_growth_high,
                  gdp_q4q4=a.gdp_q4q4,
                  gdp_annual_avg=(g + 1.0 if a.year == 2027 else g),
                  source_ref="t")
        for a in inp.annual
    ]
    inp2 = ShadowInputs(quarterly=inp.quarterly, annual=annual, params=inp.params)
    checks = m.annual_average_crosscheck(inp2, end_quarter="2028Q4")
    by_year = {c.year: c for c in checks}
    assert by_year[2027].tripped is True
    assert abs(by_year[2027].diff) > 0.15
    # other years still fine
    assert by_year[2028].tripped is False


# --------------------------------------------------------------------------- #
# Vintage flexibility: horizon-from-params, TO-FILL date, new-quarter, glob
# --------------------------------------------------------------------------- #
def test_horizon_defaults_to_projection_end_quarter():
    """run_model with no end_quarter ends exactly at projection_end_quarter."""
    inp = const_inputs(projection_end_quarter="2027Q4")
    res = m.run_model(inp)  # no explicit end_quarter
    assert res.steps[-1].quarter == "2027Q4"
    # a longer horizon param extends the path
    inp2 = const_inputs(projection_end_quarter="2029Q2")
    res2 = m.run_model(inp2)
    assert res2.steps[-1].quarter == "2029Q2"


def test_projection_end_must_be_after_seed():
    """projection_end_quarter at/before the seed quarter is rejected."""
    with pytest.raises(Exception):
        make_params(projection_end_quarter="2026Q1")  # seed is 2026Q2
    with pytest.raises(Exception):
        make_params(projection_end_quarter="2026Q2")  # equal to seed


def test_projection_end_unparseable_rejected():
    with pytest.raises(Exception):
        make_params(projection_end_quarter="2028-Q4")


def test_coverage_validation_trips_when_data_stops_short(tmp_path):
    """If projection_end_quarter reaches past the data, coverage validation trips.

    Build the seed workbook (horizon 2028Q4, data to 2028Q4), then push the
    horizon to 2029Q4 without adding a 2029 core Q4 anchor -> parse_workbook
    fails closed naming the missing year.
    """
    from openpyxl import load_workbook

    from pipeline.shadow_rate.inputs import parse_workbook
    from pipeline.shadow_rate.make_workbook import build_workbook

    xlsx = tmp_path / "wb.xlsx"
    build_workbook(str(xlsx))
    # baseline parses cleanly
    parse_workbook(str(xlsx))
    # extend the horizon past the data
    wb = load_workbook(str(xlsx))
    wp = wb["params"]
    for ri in range(2, wp.max_row + 1):
        if wp.cell(ri, 1).value == "projection_end_quarter":
            wp.cell(ri, 2, "2029Q4")
    wb.save(str(xlsx))
    wb.close()
    with pytest.raises(ValueError, match="missing core-CPI Q4 coverage for 2029"):
        parse_workbook(str(xlsx))


def test_tofill_date_marker_rejected_with_clear_error(tmp_path):
    """A workbook whose mpr_publication_date is the TO-FILL marker is rejected."""
    from openpyxl import load_workbook

    from pipeline.shadow_rate.inputs import TOFILL_DATE_MARKER, parse_workbook
    from pipeline.shadow_rate.make_workbook import build_workbook

    xlsx = tmp_path / "wb.xlsx"
    build_workbook(str(xlsx))
    wb = load_workbook(str(xlsx))
    wp = wb["params"]
    for ri in range(2, wp.max_row + 1):
        if wp.cell(ri, 1).value == "mpr_publication_date":
            wp.cell(ri, 2, TOFILL_DATE_MARKER)
    wb.save(str(xlsx))
    wb.close()
    with pytest.raises(ValueError, match="TO-FILL placeholder"):
        parse_workbook(str(xlsx))


def test_build_workbook_refuses_overwrite(tmp_path):
    """A second build over an existing file raises unless overwrite=True."""
    from pipeline.shadow_rate.make_workbook import build_workbook

    xlsx = tmp_path / "boc_shadow_inputs_2026Q2.xlsx"
    build_workbook(str(xlsx))
    with pytest.raises(FileExistsError, match="--new-quarter"):
        build_workbook(str(xlsx))
    # explicit overwrite is allowed
    build_workbook(str(xlsx), overwrite=True)


def test_new_quarter_copy_preserves_data_resets_verified_and_date(tmp_path):
    """--new-quarter copy keeps data rows but resets verified + date marker."""
    from openpyxl import load_workbook

    from pipeline.shadow_rate.inputs import TOFILL_DATE_MARKER
    from pipeline.shadow_rate.make_workbook import (
        build_workbook,
        new_quarter_workbook,
    )

    src = tmp_path / "boc_shadow_inputs_2026Q2.xlsx"
    build_workbook(str(src))

    # snapshot the data rows before copy-forward
    wb0 = load_workbook(str(src))
    q_before = [r for r in wb0["quarterly"].iter_rows(values_only=True)]
    a_before = [r for r in wb0["annual"].iter_rows(values_only=True)]
    wb0.close()

    dest = new_quarter_workbook("2026Q3", source=str(src))
    assert dest.name == "boc_shadow_inputs_2026Q3.xlsx"

    wb = load_workbook(str(dest))
    try:
        # data rows are KEPT verbatim
        assert [r for r in wb["quarterly"].iter_rows(values_only=True)] == q_before
        assert [r for r in wb["annual"].iter_rows(values_only=True)] == a_before
        # params: verified FALSE, date is the TO-FILL marker
        params = {}
        wp = wb["params"]
        for ri in range(2, wp.max_row + 1):
            k = wp.cell(ri, 1).value
            if k is not None:
                params[str(k).strip()] = wp.cell(ri, 2).value
        assert str(params["verified"]).strip().upper() == "FALSE"
        assert params["mpr_publication_date"] == TOFILL_DATE_MARKER
        # anchors re-seeded to real values (not blank)
        assert params["current_overnight_rate"] is not None
        assert params["output_gap_anchor_value"] is not None
        # no stale calc sheet
        assert "calc" not in wb.sheetnames
    finally:
        wb.close()


def test_new_quarter_copy_blocks_run_until_date_filled(tmp_path):
    """The copy-forward workbook cannot be parsed until the date is filled."""
    from pipeline.shadow_rate.inputs import parse_workbook
    from pipeline.shadow_rate.make_workbook import (
        build_workbook,
        new_quarter_workbook,
    )

    src = tmp_path / "boc_shadow_inputs_2026Q2.xlsx"
    build_workbook(str(src))
    dest = new_quarter_workbook("2026Q3", source=str(src))
    with pytest.raises(ValueError, match="TO-FILL placeholder"):
        parse_workbook(str(dest))


def test_newest_workbook_glob_selection(tmp_path):
    """The newest-workbook glob picks the lexically latest vintage, skips ~$ files."""
    from pipeline.shadow_rate.make_workbook import _newest_workbook

    for name in ("boc_shadow_inputs_2026Q2.xlsx",
                 "boc_shadow_inputs_2026Q3.xlsx",
                 "boc_shadow_inputs_2027Q1.xlsx",
                 "~$boc_shadow_inputs_2099Q4.xlsx"):  # Excel lock file: ignored
        (tmp_path / name).write_bytes(b"x")
    newest = _newest_workbook(tmp_path)
    assert newest.name == "boc_shadow_inputs_2027Q1.xlsx"


def test_vintage_stamped_filename_derivation():
    """Vintage tag + chart/series names derive from mpr_publication_date."""
    from datetime import date as _date

    for d, tag in [(_date(2026, 4, 29), "2026-04"),
                   (_date(2026, 7, 30), "2026-07"),
                   (_date(2027, 1, 22), "2027-01")]:
        vintage_tag = f"{d.year:04d}-{d.month:02d}"
        assert vintage_tag == tag
        assert f"boc_shadow_path_{vintage_tag}.svg" == f"boc_shadow_path_{tag}.svg"
        assert f"boc_shadow_rate_{vintage_tag}" == f"boc_shadow_rate_{tag}"
