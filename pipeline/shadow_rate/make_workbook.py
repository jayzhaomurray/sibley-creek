"""Author the punch-in workbook, seeded with the April 2026 MPR transcription.

Two modes:

    python -m pipeline.shadow_rate.make_workbook            # seed (April-2026)
    python -m pipeline.shadow_rate.make_workbook --overwrite # regenerate in place
    python -m pipeline.shadow_rate.make_workbook --new-quarter 2026Q3

The default seed build REFUSES to overwrite an existing workbook (its April-2026
seeds would destroy punched-in data); pass ``--overwrite`` only to deliberately
regenerate the seed layout. The quarterly refresh workflow is ``--new-quarter``:
it copies the newest workbook forward to a fresh quarter, resetting ``verified``
to FALSE and ``mpr_publication_date`` to a TO-FILL marker the parser rejects,
re-seeding the gap anchor + overnight rate from the data tails, and KEEPING all
quarterly/annual data rows so Jay edits last quarter's numbers in place.

The workbook ships with ``verified=FALSE``: the runner refuses to emit a real
artifact until Jay has checked every transcribed cell against the MPR PDF and
flipped that flag. Every data field carries a ``source_ref`` provenance string.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).parents[2]
DEFAULT_OUT = (
    PROJECT_ROOT / "work" / "research" / "shadow_rate" / "boc_shadow_inputs_2026Q2.xlsx"
)

# Source references (transcribed from Jay's screenshots of the April 2026 MPR).
REF_TABLE3 = "MPR Apr-2026, Table 3"
REF_TABLE2 = "MPR Apr-2026, Table 2"
REF_TR119 = "BoC Technical Report 119, Table 2.3"
REF_ELB = "BoC effective lower bound statements, 2009 & Apr 2020"
REF_NEUTRAL = (
    "MPR Apr-2026, Appendix 'Potential output and the nominal neutral rate' "
    "(neutral rate 2.25-3.25%, unchanged from Apr-2025); "
    "https://www.bankofcanada.ca/publications/mpr/mpr-2026-04-29/appendix/"
)
REF_OVERNIGHT = "data/processed/overnight_rate_target.csv (tail)"
REF_MPRDATE = "https://www.bankofcanada.ca/publications/mpr/mpr-2026-04-29/"
REF_GAP_ANCHOR = (
    "BoC Valet INDINF_OUTGAPMPR_Q (staff output gap, current MPR vintage), "
    "last published observation"
)

# quarterly sheet rows: (quarter, core_cpi, total_cpi_ref, gdp_qq_ann, anchor_type, ref)
QUARTERLY_ROWS = [
    ("2025Q3", 3.1, 2.0, 2.4, "quarterly", REF_TABLE3),
    ("2025Q4", 2.8, 2.2, -0.6, "quarterly", REF_TABLE3),
    ("2026Q1", 2.4, 2.2, 1.5, "quarterly", REF_TABLE3),
    ("2026Q2", 2.1, 2.6, 1.5, "quarterly", REF_TABLE3),
    ("2026Q4", 2.0, 2.2, None, "q4q4", REF_TABLE3),
    ("2027Q4", 2.2, 2.0, None, "q4q4", REF_TABLE3),
    ("2028Q4", 2.0, 2.0, None, "q4q4", REF_TABLE3),
]

# annual sheet rows: (year, pot_low, pot_high, gdp_q4q4, gdp_annual_avg, ref)
# Potential ranges from Table 2; Q4/Q4 real GDP from Table 3; annual-AVERAGE
# real GDP growth from Table 2 (2026=1.2, 2027=1.6, 2028=1.7).
ANNUAL_ROWS = [
    (2025, 2.3, 2.3, None, None, f"{REF_TABLE2} (point estimate)"),
    (2026, 0.8, 1.6, 1.8, 1.2,
     f"{REF_TABLE2} potential + annual-avg GDP; {REF_TABLE3} Q4/Q4 GDP"),
    (2027, 0.8, 1.8, 1.4, 1.6,
     f"{REF_TABLE2} potential + annual-avg GDP; {REF_TABLE3} Q4/Q4 GDP"),
    (2028, 1.0, 2.0, 1.9, 1.7,
     f"{REF_TABLE2} potential + annual-avg GDP; {REF_TABLE3} Q4/Q4 GDP"),
]

# params sheet rows: (key, value, source_ref)
# current_overnight_rate and the output-gap anchor quarter/value are filled at
# build time (overnight from the rate CSV tail; the anchor from the last row of
# data/raw/output_gap_mpr.csv, so quarterly refreshes pick up the new vintage).
PARAM_ROWS_TEMPLATE = [
    ("mpr_publication_date", "2026-04-29", REF_MPRDATE),
    ("projection_end_quarter", "2028Q4", "MPR Apr-2026 projection horizon"),
    ("current_overnight_rate", None, REF_OVERNIGHT),  # filled at runtime
    ("output_gap_anchor_quarter", None, REF_GAP_ANCHOR),  # filled at runtime
    ("output_gap_anchor_value", None, REF_GAP_ANCHOR),    # filled at runtime
    ("neutral_range_low", 2.25, REF_NEUTRAL),
    ("neutral_range_high", 3.25, REF_NEUTRAL),
    ("rho", 0.85, REF_TR119),
    ("phi_pi", 4.65, REF_TR119),
    ("phi_gap", 0.40, REF_TR119),
    ("inflation_target", 2.0, "BoC 2% CPI inflation target"),
    ("inflation_converge_quarters", 4, "TR-119 rule horizon (t+4)"),
    ("elb_floor", 0.25, REF_ELB),
    ("verified", "FALSE", "Jay flips to TRUE after checking every cell vs MPR PDF"),
]


_HEADER_FILL = PatternFill("solid", fgColor="222222")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_NOTE_FONT = Font(italic=True, color="888888")


def _read_last_overnight_rate() -> float:
    """Read the last value from data/processed/overnight_rate_target.csv."""
    csv = PROJECT_ROOT / "data" / "processed" / "overnight_rate_target.csv"
    last_val = None
    with open(csv, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.lower().startswith("date"):
                continue
            parts = line.split(",")
            if len(parts) >= 2 and parts[1].strip():
                last_val = float(parts[1].strip())
    if last_val is None:
        raise ValueError(f"could not read a value from {csv}")
    return last_val


def _date_to_quarter(iso: str) -> str:
    """ISO date string 'YYYY-MM-DD' -> 'YYYYQn' calendar quarter."""
    year = int(iso[:4])
    month = int(iso[5:7])
    qn = (month - 1) // 3 + 1
    return f"{year}Q{qn}"


def _read_output_gap_anchor() -> tuple[str, float]:
    """Read the last (date, value) from data/raw/output_gap_mpr.csv.

    Returns the anchor as ('YYYYQn', value). Reading the raw CSV's last row means
    a quarterly refresh (which rewrites that file from Valet) automatically picks
    up a newer published staff output-gap vintage.
    """
    csv = PROJECT_ROOT / "data" / "raw" / "output_gap_mpr.csv"
    last_date = None
    last_val = None
    with open(csv, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.lower().startswith("date"):
                continue
            parts = line.split(",")
            if len(parts) >= 2 and parts[0].strip() and parts[1].strip():
                last_date = parts[0].strip()
                last_val = float(parts[1].strip())
    if last_date is None or last_val is None:
        raise ValueError(f"could not read an observation from {csv}")
    return _date_to_quarter(last_date), last_val


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="left")


def _autosize(ws, widths: dict[int, int]) -> None:
    from openpyxl.utils import get_column_letter
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_workbook(out_path: str | Path = DEFAULT_OUT, *, overwrite: bool = False) -> Path:
    """Build and save the seeded workbook. Returns the output path.

    Refuses to overwrite an existing workbook unless ``overwrite=True``: the
    seeds here are the April-2026 vintage, so regenerating on top of a punched-in
    later vintage would destroy Jay's transcribed MPR data. The quarterly refresh
    workflow is copy-forward (``--new-quarter``), not regenerate.
    """
    out_path = Path(out_path)
    if out_path.exists() and not overwrite:
        raise FileExistsError(
            f"workbook already exists: {out_path}\n"
            f"  build_workbook seeds the April-2026 vintage and would DESTROY any "
            f"punched-in data. To start a NEW quarter, copy forward instead:\n"
            f"    python -m pipeline.shadow_rate.make_workbook --new-quarter <YYYYQn>\n"
            f"  To deliberately regenerate the seed layout in place, pass --overwrite."
        )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    overnight = _read_last_overnight_rate()
    anchor_quarter, anchor_value = _read_output_gap_anchor()

    wb = Workbook()

    # --- quarterly sheet ---
    ws = wb.active
    ws.title = "quarterly"
    headers = [
        "quarter",
        "core_cpi_yoy_forecast",
        "total_cpi_yoy_reference",
        "gdp_growth_qq_ann_forecast",
        "anchor_type",
        "source_ref",
    ]
    ws.append(headers)
    for q, core, total, gdp, atype, ref in QUARTERLY_ROWS:
        ws.append([q, core, total, gdp, atype, ref])
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
        for c in row:
            if c.value is not None:
                c.number_format = "0.0"
    _style_header(ws, len(headers))
    _autosize(ws, {1: 10, 2: 22, 3: 22, 4: 26, 5: 12, 6: 30})
    ws.freeze_panes = "A2"

    # --- annual sheet ---
    wa = wb.create_sheet("annual")
    headers_a = [
        "year",
        "potential_growth_low",
        "potential_growth_high",
        "gdp_q4q4",
        "gdp_annual_avg",
        "source_ref",
    ]
    wa.append(headers_a)
    for yr, lo, hi, gdp, gdp_avg, ref in ANNUAL_ROWS:
        wa.append([yr, lo, hi, gdp, gdp_avg, ref])
    for row in wa.iter_rows(min_row=2, min_col=2, max_col=5):
        for c in row:
            if c.value is not None:
                c.number_format = "0.0"
    _style_header(wa, len(headers_a))
    _autosize(wa, {1: 8, 2: 22, 3: 22, 4: 12, 5: 16, 6: 50})
    wa.freeze_panes = "A2"

    # --- params sheet ---
    wp = wb.create_sheet("params")
    headers_p = ["key", "value", "source_ref"]
    wp.append(headers_p)
    for key, val, ref in PARAM_ROWS_TEMPLATE:
        if key == "current_overnight_rate":
            val = overnight
        elif key == "output_gap_anchor_quarter":
            val = anchor_quarter
        elif key == "output_gap_anchor_value":
            val = anchor_value
        wp.append([key, val, ref])
    _style_header(wp, len(headers_p))
    _autosize(wp, {1: 30, 2: 16, 3: 70})
    wp.freeze_panes = "A2"

    # A small operating note below the params block.
    note_row = wp.max_row + 2
    wp.cell(row=note_row, column=1,
            value="NOTE: verified must be TRUE for a real run. "
                  "Use --force-unverified only for watermarked drafts.").font = _NOTE_FONT

    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
# Quarterly copy-forward (--new-quarter)
# --------------------------------------------------------------------------- #
# Marker the parser rejects until Jay fills in the real MPR date. Kept in sync
# with inputs.TOFILL_DATE_MARKER (imported here so there is one definition).
from pipeline.shadow_rate.inputs import TOFILL_DATE_MARKER  # noqa: E402

_SOURCE_REF_REMINDER = (
    "TO-VERIFY vs new MPR: update value AND this source_ref to the new vintage"
)


def _newest_workbook(folder: Path) -> Path:
    """Newest boc_shadow_inputs_<YYYY>Q<n>.xlsx in folder (lexically last)."""
    candidates = sorted(
        p for p in folder.glob("boc_shadow_inputs_*.xlsx")
        if not p.name.startswith("~$")
    )
    if not candidates:
        raise FileNotFoundError(
            f"no boc_shadow_inputs_*.xlsx workbook found in {folder}"
        )
    return candidates[-1]


def new_quarter_workbook(
    new_quarter: str,
    *,
    source: str | Path | None = None,
    out_path: str | Path | None = None,
) -> Path:
    """Copy the newest workbook forward to a fresh quarter for punch-in.

    Reads the existing newest workbook (or ``source`` if given), copies it to
    ``boc_shadow_inputs_<new_quarter>.xlsx`` and resets it for a new MPR:

      - ``verified`` -> FALSE (Jay must re-check every cell);
      - ``mpr_publication_date`` -> a TO-FILL marker the parser rejects with a
        clear message, so a run can't accidentally reuse the stale date;
      - gap anchor (quarter + value) re-seeded from the tail of
        ``data/raw/output_gap_mpr.csv``;
      - ``current_overnight_rate`` re-seeded from the tail of
        ``data/processed/overnight_rate_target.csv``;
      - ALL quarterly / annual data rows KEPT (Jay edits last quarter's numbers
        in place — faster punch-in than blank cells; the verified gate guards
        against forgetting);
      - a source_ref reminder note appended to the params sheet.

    The ``calc`` sheet, if present, is dropped — it is regenerated on the next
    run. Returns the new workbook path.
    """
    from openpyxl import load_workbook

    # validate the requested quarter label
    nq = str(new_quarter).strip().upper()
    if (len(nq) != 6 or nq[4] != "Q" or not nq[:4].isdigit() or nq[5] not in "1234"):
        raise ValueError(f"--new-quarter must look like '2026Q3', got {new_quarter!r}")
    nq = nq[:4] + "Q" + nq[5]

    src = Path(source) if source is not None else _newest_workbook(DEFAULT_OUT.parent)
    dest = (
        Path(out_path) if out_path is not None
        else src.with_name(f"boc_shadow_inputs_{nq}.xlsx")
    )
    if dest.exists():
        raise FileExistsError(
            f"target workbook already exists: {dest}\n"
            f"  refusing to overwrite an existing {nq} workbook (it may hold "
            f"punched-in data). Move or delete it first if you really want a "
            f"fresh copy-forward."
        )

    overnight = _read_last_overnight_rate()
    anchor_quarter, anchor_value = _read_output_gap_anchor()

    wb = load_workbook(src)  # full load, preserve everything
    # Drop any stale calc sheet; it regenerates on the next run.
    if "calc" in wb.sheetnames:
        del wb["calc"]
    if "output" in wb.sheetnames:
        del wb["output"]

    wp = wb["params"]
    # Map params key -> row (column A = key, B = value, C = source_ref).
    key_row: dict[str, int] = {}
    for ri in range(2, wp.max_row + 1):
        k = wp.cell(ri, 1).value
        if k is not None:
            key_row[str(k).strip()] = ri

    def _set(key: str, value, source_ref=None):
        ri = key_row.get(key)
        if ri is None:
            raise KeyError(f"params sheet missing key {key!r} in {src}")
        wp.cell(ri, 2, value)
        if source_ref is not None:
            wp.cell(ri, 3, source_ref)

    _set("verified", "FALSE")
    _set("mpr_publication_date", TOFILL_DATE_MARKER, _SOURCE_REF_REMINDER)
    _set("output_gap_anchor_quarter", anchor_quarter)
    _set("output_gap_anchor_value", anchor_value)
    _set("current_overnight_rate", overnight)
    # Nudge Jay to re-source the headline MPR transcription fields.
    if "projection_end_quarter" in key_row:
        _set("projection_end_quarter", wp.cell(key_row["projection_end_quarter"], 2).value,
             _SOURCE_REF_REMINDER)

    # Append a fresh source_ref reminder note below the params block.
    note_row = wp.max_row + 2
    wp.cell(
        row=note_row, column=1,
        value=(f"NEW-QUARTER COPY for {nq}: data rows carry LAST quarter's MPR "
               f"numbers — edit them in place against the new MPR Tables 2-3, set "
               f"projection_end_quarter + mpr_publication_date, then flip verified "
               f"to TRUE. The run is blocked until the TO-FILL date is replaced."),
    ).font = _NOTE_FONT

    wb.save(dest)
    wb.close()
    return dest


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build the shadow-rate punch-in workbook")
    parser.add_argument("--out", default=str(DEFAULT_OUT), help="output xlsx path")
    parser.add_argument("--overwrite", action="store_true",
                        help="allow regenerating the seed layout over an existing file")
    parser.add_argument("--new-quarter", default=None, metavar="YYYYQn",
                        help="copy the newest workbook forward to a fresh quarter "
                             "for punch-in (resets verified/date, re-seeds anchors, "
                             "keeps data rows)")
    args = parser.parse_args()

    if args.new_quarter:
        path = new_quarter_workbook(args.new_quarter)
        print(f"Wrote new-quarter workbook: {path}")
        _aq, _av = _read_output_gap_anchor()
        print(f"  reset verified=FALSE; mpr_publication_date -> TO-FILL marker "
              f"(run blocked until filled)")
        print(f"  re-seeded current_overnight_rate from CSV tail: "
              f"{_read_last_overnight_rate()}")
        print(f"  re-seeded output-gap anchor from "
              f"data/raw/output_gap_mpr.csv tail: {_aq} = {_av}")
        print(f"  kept all quarterly/annual data rows (edit in place vs the new MPR)")
    else:
        path = build_workbook(args.out, overwrite=args.overwrite)
        print(f"Wrote workbook: {path}")
        print(f"  seeded current_overnight_rate from CSV tail: "
              f"{_read_last_overnight_rate()}")
        _aq, _av = _read_output_gap_anchor()
        print(f"  seeded output-gap anchor from data/raw/output_gap_mpr.csv tail: "
              f"{_aq} = {_av}")
