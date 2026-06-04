"""Write a live-formula ``calc`` sheet back into the punch-in workbook.

After the model runs, this opens the same workbook Jay edits and writes (or
replaces) a sheet named ``calc`` placed FIRST in the sheet order, so it is what
Jay sees when he opens the file. The sheet is a dense quarterly grid that
reproduces the entire policy-rule path with **live Excel formulas** referencing
the input sheets (``quarterly`` / ``annual`` / ``params``). Change an input and
the whole path recomputes in Excel.

The single source of truth for *agreement* is the tested Python engine
(``pipeline/shadow_rate/model.py``). Each formula column is paired with a static
``... (python)`` column carrying the engine's value from the last run, and a
``diff`` column = ABS(formula - python) with red conditional formatting if the
divergence exceeds a tolerance. That is the audit handshake: Excel recomputes
everything live, and the diff columns prove the live path matches the engine.

The input sheets (``quarterly`` / ``annual`` / ``params``) are never touched.

Windows file-lock handling: if the workbook is open in Excel, openpyxl's save
raises ``PermissionError``. We catch it and write a companion file
``boc_shadow_output_2026Q2.xlsx`` in the same folder instead, returning the path
written and whether it was the companion fallback.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pipeline.shadow_rate.model import (
    ShadowResult,
    build_core_cpi_path,
    ord_to_quarter,
    quarter_to_ord,
)

SHEET_NAME = "calc"

# Diff tolerance for the audit handshake (red fill above this). The grid runs
# from the gap anchor quarter through the model horizon plus the
# inflation-convergence headroom, so terminal-quarter t+4 lookups land on real
# rows; both bounds are derived at build time from the params/result.
DIFF_TOL = 0.0005

# --- styling primitives (Vignelli-ish restraint: thin grey rules, no fill noise) ---
_INK = "1A1A1A"
_GREY = "BFBFBF"
_HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
_RED_FILL = PatternFill("solid", fgColor="C0392B")
_GREY_FILL = PatternFill("solid", fgColor="EDEDED")
_PY_FILL = PatternFill("solid", fgColor="FBF7E8")  # faint cream: static engine values
_THIN = Side(style="thin", color=_GREY)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_NUM3 = "0.000"
_NUM2 = "0.00"
_NUM4 = "0.0000"

# Input-sheet cell geometry (header row 1, data from row 2).
# quarterly: A quarter | B core | C total | D gdp  (1-indexed -> letters)
_Q_CORE_COL = "B"
_Q_GDP_COL = "D"
# annual: A year | B pot_low | C pot_high | D gdp_q4q4
_A_LOW_COL = "B"
_A_HIGH_COL = "C"
_A_GDPQ4_COL = "D"


# Calc-sheet column layout (1-indexed). Keep in sync with _COLS below.
@dataclass(frozen=True)
class _Col:
    idx: int
    header: str
    fmt: str | None


_COLS = [
    _Col(1, "quarter", None),
    _Col(2, "core CPI y/y (%)", _NUM3),
    _Col(3, "gdp q/q ann (%)", _NUM2),
    _Col(4, "potential (%)", _NUM2),
    _Col(5, "output gap (pp)", _NUM3),
    _Col(6, "gap (python)", _NUM3),
    _Col(7, "gap diff", _NUM4),
    _Col(8, "pi t+4 (%)", _NUM3),
    _Col(9, "neutral mid (%)", _NUM3),
    _Col(10, "infl term", _NUM3),
    _Col(11, "gap term", _NUM3),
    _Col(12, "bracket", _NUM3),
    _Col(13, "shadow rate R (%)", _NUM3),
    _Col(14, "R (python)", _NUM3),
    _Col(15, "R diff", _NUM4),
]
_NCOLS = len(_COLS)

# Convenience: column letters.
COL_QUARTER = get_column_letter(1)
COL_CORE = get_column_letter(2)
COL_GDP = get_column_letter(3)
COL_POT = get_column_letter(4)
COL_GAP = get_column_letter(5)
COL_GAP_PY = get_column_letter(6)
COL_GAP_DIFF = get_column_letter(7)
COL_PIT4 = get_column_letter(8)
COL_NEUTRAL = get_column_letter(9)
COL_INFL = get_column_letter(10)
COL_GAPTERM = get_column_letter(11)
COL_BRACKET = get_column_letter(12)
COL_RATE = get_column_letter(13)
COL_RATE_PY = get_column_letter(14)
COL_RATE_DIFF = get_column_letter(15)


@dataclass
class OutputWriteResult:
    path: Path
    used_companion: bool


# --------------------------------------------------------------------------- #
# Reference helpers — map a quarter ordinal to a row in an input sheet.
# --------------------------------------------------------------------------- #
def _interp_formula(o: int, known_rows: dict[int, int]) -> str:
    """Live core-CPI formula for quarter ordinal ``o``.

    - direct reference if ``o`` is a known quarterly point;
    - linear interpolation between the bracketing known points otherwise;
    - hold at the last known point past the final anchor.
    """
    ks = sorted(known_rows)
    first, last = ks[0], ks[-1]
    if o in known_rows:
        return f"=quarterly!{_Q_CORE_COL}{known_rows[o]}"
    if o > last:
        return f"=quarterly!{_Q_CORE_COL}{known_rows[last]}"
    if o < first:
        # Should never happen: grid starts at the anchor, anchor >= first known.
        return f"=quarterly!{_Q_CORE_COL}{known_rows[first]}"
    lo = max(k for k in ks if k < o)
    hi = min(k for k in ks if k > o)
    span = hi - lo
    offset = o - lo
    lo_ref = f"quarterly!{_Q_CORE_COL}{known_rows[lo]}"
    hi_ref = f"quarterly!{_Q_CORE_COL}{known_rows[hi]}"
    return f"={lo_ref}+({offset}/{span})*({hi_ref}-{lo_ref})"


def _gdp_formula(o: int, gdp_rows: dict[int, int], year_rows: dict[int, int]) -> str:
    """Live GDP q/q-ann formula: direct quarterly cell where present, else that
    year's gdp_q4q4 from the annual sheet."""
    if o in gdp_rows:
        return f"=quarterly!{_Q_GDP_COL}{gdp_rows[o]}"
    yr = o // 4
    return f"=annual!{_A_GDPQ4_COL}{year_rows[yr]}"


def _potential_formula(o: int, year_rows: dict[int, int]) -> str:
    """Potential = midpoint of that year's low/high range in the annual sheet."""
    yr = o // 4
    row = year_rows[yr]
    return f"=(annual!{_A_LOW_COL}{row}+annual!{_A_HIGH_COL}{row})/2"


# --------------------------------------------------------------------------- #
# Sheet builder
# --------------------------------------------------------------------------- #
def _build_calc_ws(wb, res: ShadowResult, p, inp) -> None:
    """Create/replace the ``calc`` sheet in wb (placed first)."""
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    # Remove any legacy values-only "output" sheet from earlier versions so the
    # final order is exactly ['calc', 'quarterly', 'annual', 'params'].
    if "output" in wb.sheetnames:
        del wb["output"]
    ws = wb.create_sheet(SHEET_NAME, index=0)

    bold = Font(bold=True, color=_INK)
    title_font = Font(bold=True, size=14, color=_INK)
    note_font = Font(italic=True, size=9, color="595959")

    draft = not p.verified
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # --- input-sheet row maps (built from the parsed inputs, in sheet order) ---
    # quarterly rows: data starts at row 2 in input order.
    core_known_rows: dict[int, int] = {}
    gdp_rows: dict[int, int] = {}
    for i, qr in enumerate(inp.quarterly):
        row = i + 2
        o = quarter_to_ord(qr.quarter)
        core_known_rows[o] = row
        if qr.gdp_growth_qq_ann_forecast is not None:
            gdp_rows[o] = row
    year_rows: dict[int, int] = {a.year: i + 2 for i, a in enumerate(inp.annual)}

    # --- params cell map (key -> row in params sheet) for absolute references ---
    param_rows: dict[str, int] = {}
    wp = wb["params"]
    for ridx, rowvals in enumerate(wp.iter_rows(values_only=True), start=1):
        if ridx == 1:
            continue
        if rowvals and rowvals[0] is not None:
            param_rows[str(rowvals[0]).strip()] = ridx

    def pref(key: str) -> str:
        """Absolute reference to a params value cell (column B)."""
        return f"params!$B${param_rows[key]}"

    # --- grid geometry ---
    anchor_ord = quarter_to_ord(p.output_gap_anchor_quarter)
    seed_ord = quarter_to_ord(res.seed_quarter)
    end_ord = quarter_to_ord(res.steps[-1].quarter)
    grid_end_ord = end_ord + p.inflation_converge_quarters
    grid_start_ord = anchor_ord

    # Engine static values, keyed by ordinal, for the python-check columns.
    gap_py = dict(res.gap_path)  # anchor..end_ord
    rate_py = {quarter_to_ord(s.quarter): s.rate for s in res.steps}  # seed..end_ord
    core_full = build_core_cpi_path(inp, grid_end_ord)

    # ------------------------------------------------------------------ #
    # Header / provenance block
    # ------------------------------------------------------------------ #
    r = 1
    ws.cell(r, 1, "BoC Shadow Policy Rate — calc (live Excel formulas)").font = title_font
    r += 1
    ws.cell(r, 1, f"Run: {now}").font = Font(size=9, color="595959")
    r += 1

    status_cell = ws.cell(r, 1)
    if draft:
        status_cell.value = "UNVERIFIED DRAFT"
        status_cell.font = Font(bold=True, size=14, color="FFFFFF")
        status_cell.fill = _RED_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    else:
        status_cell.value = "VERIFIED"
        status_cell.font = Font(bold=True, size=12, color="2E7D32")
    r += 2

    block = [
        ("Anchor provenance",
         f"output gap anchored at {p.output_gap_anchor_quarter} = "
         f"{p.output_gap_anchor_value:+.2f}pp (source: BoC Valet "
         f"INDINF_OUTGAPMPR_Q, staff output gap, last published obs; "
         f"rolled forward to seed via live formulas)"),
        ("Seed quarter / seed rate",
         f"{res.seed_quarter} @ {res.seed_rate:.2f}% (actual overnight rate at MPR)"),
        ("R*_nom (nominal neutral midpoint)",
         f"{p.neutral_nominal_mid:.2f}% "
         f"(range {p.neutral_range_low:.2f}-{p.neutral_range_high:.2f})"),
        ("Rule citation",
         f"ToTEM III, TR-119 Table 2.3: rho={p.rho}, phi_pi={p.phi_pi}, "
         f"phi_gap={p.phi_gap}; inflation target {p.inflation_target:.1f}%, "
         f"ELB floor {p.elb_floor:.2f}%, t+{p.inflation_converge_quarters} lookup"),
        ("MPR publication date", p.mpr_publication_date.isoformat()),
    ]
    for label, val in block:
        ws.cell(r, 1, label).font = bold
        ws.cell(r, 2, val)
        r += 1

    # The audit-handshake explainer line.
    explain = (
        "All white cells are live formulas — change an input and the path "
        "recomputes. The 'python' columns are the engine's values from the last "
        "run; diff columns flag any divergence (red if > 0.0005)."
    )
    exp_cell = ws.cell(r, 1, explain)
    exp_cell.font = note_font
    exp_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_NCOLS)
    r += 2

    # ------------------------------------------------------------------ #
    # Grid header row
    # ------------------------------------------------------------------ #
    header_row = r
    for c in _COLS:
        cell = ws.cell(header_row, c.idx, c.header)
        cell.font = bold
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    r += 1

    first_data_row = r
    row_of_ord: dict[int, int] = {}

    for o in range(grid_start_ord, grid_end_ord + 1):
        row = r
        row_of_ord[o] = row
        is_rate_row = seed_ord <= o <= end_ord
        is_gap_row = anchor_ord <= o <= end_ord

        # A: quarter label
        qa = ws.cell(row, 1, ord_to_quarter(o))
        qa.alignment = Alignment(horizontal="center")

        # B: core CPI y/y (live)
        ws.cell(row, 2, _interp_formula(o, core_known_rows))

        # C/D/E: gdp, potential, gap — only through the model horizon (end_ord).
        # The trailing headroom rows (end_ord+1 .. grid_end) exist solely to host
        # core-CPI cells for the terminal quarters' t+4 lookups; the annual sheet
        # has no rows past end_ord's year, so gdp/potential/gap stop here.
        if o <= end_ord:
            # C: gdp q/q ann (live)
            ws.cell(row, 3, _gdp_formula(o, gdp_rows, year_rows))
            # D: potential (live)
            ws.cell(row, 4, _potential_formula(o, year_rows))
            # E: output gap (live). Anchor row = params anchor value; later rows
            # = gap above + (gdp - potential)/4 (t+1 timing, matches engine).
            if o == anchor_ord:
                ws.cell(row, 5, f"={pref('output_gap_anchor_value')}")
            else:
                prev = row - 1
                ws.cell(
                    row, 5,
                    f"={COL_GAP}{prev}+({COL_GDP}{row}-{COL_POT}{row})/4",
                )

        # F/G: gap python + diff (only where engine produced a gap)
        if is_gap_row and o in gap_py:
            pyc = ws.cell(row, 6, round(gap_py[o], 6))
            pyc.fill = _PY_FILL
            ws.cell(row, 7, f"=ABS({COL_GAP}{row}-{COL_GAP_PY}{row})")

        # H: pi t+4 (live) = core CPI cell `converge` rows below. The grid runs
        # to end_ord + converge, so every rate row (o <= end_ord) has its t+4
        # row on the grid; the trailing headroom rows themselves do not.
        if o + p.inflation_converge_quarters <= grid_end_ord:
            t4_row = row + p.inflation_converge_quarters
            ws.cell(row, 8, f"={COL_CORE}{t4_row}")

        # Rate decomposition columns — only for rate rows (seed..end).
        if is_rate_row:
            # I: neutral mid = (neutral_low + neutral_high)/2
            ws.cell(
                row, 9,
                f"=({pref('neutral_range_low')}+{pref('neutral_range_high')})/2",
            )
            # J: infl term = phi_pi*(pi_t4 - target)
            ws.cell(
                row, 10,
                f"={pref('phi_pi')}*({COL_PIT4}{row}-{pref('inflation_target')})",
            )
            # K: gap term = phi_gap*gap
            ws.cell(row, 11, f"={pref('phi_gap')}*{COL_GAP}{row}")
            # L: bracket = neutral_mid + infl + gap
            ws.cell(
                row, 12,
                f"={COL_NEUTRAL}{row}+{COL_INFL}{row}+{COL_GAPTERM}{row}",
            )
            # M: shadow rate R. Seed row = current overnight; later rows =
            # MAX(rho*R_above + (1-rho)*bracket_above, elb_floor).
            if o == seed_ord:
                ws.cell(row, 13, f"={pref('current_overnight_rate')}")
            else:
                prev = row - 1
                ws.cell(
                    row, 13,
                    f"=MAX({pref('rho')}*{COL_RATE}{prev}"
                    f"+(1-{pref('rho')})*{COL_BRACKET}{prev},{pref('elb_floor')})",
                )
            # N/O: rate python + diff
            if o in rate_py:
                pyc = ws.cell(row, 14, round(rate_py[o], 6))
                pyc.fill = _PY_FILL
                ws.cell(row, 15, f"=ABS({COL_RATE}{row}-{COL_RATE_PY}{row})")

        # Pre-seed roll-forward region: grey the rate-block cells (no iteration
        # before the seed) so it reads as not-applicable.
        if o < seed_ord:
            for cidx in range(9, _NCOLS + 1):
                ws.cell(row, cidx).fill = _GREY_FILL

        # number formats + borders for the whole row
        for c in _COLS:
            cell = ws.cell(row, c.idx)
            cell.border = _BORDER
            if c.fmt and cell.value is not None:
                cell.number_format = c.fmt

        r += 1

    last_data_row = r - 1

    # ------------------------------------------------------------------ #
    # Conditional formatting on the diff columns (red if > tolerance)
    # ------------------------------------------------------------------ #
    red_text_fill = PatternFill("solid", fgColor="F4CCCC")
    red_font = Font(color="990000", bold=True)
    for col in (COL_GAP_DIFF, COL_RATE_DIFF):
        rng = f"{col}{first_data_row}:{col}{last_data_row}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=[str(DIFF_TOL)],
                       fill=red_text_fill, font=red_font),
        )

    # ------------------------------------------------------------------ #
    # Footer note
    # ------------------------------------------------------------------ #
    note_r = last_data_row + 2
    note = (
        "Engine: pipeline/shadow_rate/model.py — single source of truth for "
        "agreement. White cells are live Excel formulas referencing the "
        "quarterly / annual / params sheets; the 'python' columns hold the "
        "engine's last-run values and the diff columns (= ABS(formula-python)) "
        "go red if any divergence exceeds 0.0005. Grey rate cells in the "
        "pre-seed roll-forward region (anchor through the quarter before seed) "
        "are intentionally blank — the rule does not iterate before the seed. "
        "Methodology: "
        "claude-ref/research/shadow_rate/shadow_rate_methodology_2026-04.md"
    )
    note_cell = ws.cell(note_r, 1, note)
    note_cell.font = note_font
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=note_r, start_column=1,
                   end_row=note_r, end_column=_NCOLS)

    # ------------------------------------------------------------------ #
    # Column widths + freeze
    # ------------------------------------------------------------------ #
    widths = [9, 13, 13, 11, 13, 12, 9, 11, 12, 11, 11, 11, 14, 12, 9]
    for c, w in zip(_COLS, widths):
        ws.column_dimensions[get_column_letter(c.idx)].width = w
    ws.freeze_panes = ws.cell(first_data_row, 1)


def write_output_sheet(xlsx_path: str | Path, res: ShadowResult, p) -> OutputWriteResult:
    """Write/replace the ``calc`` sheet in the workbook at ``xlsx_path``.

    Preserves the input sheets. On a Windows file lock (workbook open in Excel),
    falls back to a companion file ``boc_shadow_output_<stem-suffix>.xlsx`` in the
    same folder and returns ``used_companion=True``.

    Note: ``p`` is accepted for signature stability with run.py, but the inputs
    are re-parsed from the workbook so the formula references line up exactly
    with the live sheet rows (input order is the source of truth for cell refs).
    """
    from pipeline.shadow_rate.inputs import parse_workbook

    xlsx_path = Path(xlsx_path)
    inp = parse_workbook(xlsx_path)
    wb = load_workbook(xlsx_path)  # full load (not read_only) so we can save back
    try:
        _build_calc_ws(wb, res, p, inp)
        try:
            wb.save(xlsx_path)
            return OutputWriteResult(path=xlsx_path, used_companion=False)
        except PermissionError:
            companion = xlsx_path.with_name(
                xlsx_path.name.replace("_inputs_", "_output_")
                if "_inputs_" in xlsx_path.name
                else f"{xlsx_path.stem}_output{xlsx_path.suffix}"
            )
            wb.save(companion)
            return OutputWriteResult(path=companion, used_companion=True)
    finally:
        wb.close()
