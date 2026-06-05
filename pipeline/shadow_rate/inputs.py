"""Workbook parsing + fail-closed validation for the shadow-rate tool.

Reads the three-sheet punch-in workbook (quarterly / annual / params) with
openpyxl and validates every field with pydantic. Validation is fail-closed:
out-of-bounds or missing values raise rather than silently defaulting, so a
fat-fingered transcription in Jay's spreadsheet stops the run instead of
producing a plausible-but-wrong path.

Sheet schema (each sheet carries a ``source_ref`` provenance column):

  quarterly: quarter, core_cpi_yoy_forecast, total_cpi_yoy_reference,
             gdp_growth_qq_ann_forecast, anchor_type, source_ref
  annual:    year, potential_growth_low, potential_growth_high, gdp_q4q4,
             source_ref
  params:    key/value rows (mpr_publication_date, current_overnight_rate,
             output_gap_anchor_quarter, output_gap_anchor_value,
             neutral_range_low/high, rho, phi_pi,
             phi_gap, inflation_target,
             elb_floor, verified) each with a source_ref
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator, model_validator


# --------------------------------------------------------------------------- #
# Validated row / param models
# --------------------------------------------------------------------------- #
class QuarterlyRow(BaseModel):
    """One row of the quarterly sheet.

    ``quarter`` is a string like "2026Q2". ``core_cpi_yoy_forecast`` is the
    model input (average of CPI-trim and CPI-median, per MPR Table 3 footnote).
    ``total_cpi_yoy_reference`` is reference-only and NOT consumed by the model.
    ``anchor_type`` is "quarterly" for near-term observed rows or "q4q4" for the
    Q4/Q4 anchor rows the model interpolates between.
    """

    quarter: str
    core_cpi_yoy_forecast: float = Field(..., ge=-5.0, le=15.0)
    total_cpi_yoy_reference: Optional[float] = Field(None, ge=-5.0, le=15.0)
    gdp_growth_qq_ann_forecast: Optional[float] = Field(None, ge=-30.0, le=30.0)
    anchor_type: str
    source_ref: str

    @field_validator("quarter")
    @classmethod
    def _check_quarter(cls, v: str) -> str:
        v = v.strip()
        if len(v) != 6 or v[4].upper() != "Q" or not v[:4].isdigit() or v[5] not in "1234":
            raise ValueError(f"quarter must look like '2026Q2', got {v!r}")
        return v[:4] + "Q" + v[5]

    @field_validator("anchor_type")
    @classmethod
    def _check_anchor(cls, v: str) -> str:
        v = v.strip().lower()
        if v not in ("quarterly", "q4q4"):
            raise ValueError(f"anchor_type must be 'quarterly' or 'q4q4', got {v!r}")
        return v

    @field_validator("source_ref")
    @classmethod
    def _check_ref(cls, v: str) -> str:
        if not v or not v.strip():
            raise ValueError("source_ref must not be empty")
        return v.strip()


class AnnualRow(BaseModel):
    """One row of the annual sheet (potential-growth ranges + Q4/Q4 GDP)."""

    year: int = Field(..., ge=2000, le=2100)
    potential_growth_low: float = Field(..., ge=-5.0, le=10.0)
    potential_growth_high: float = Field(..., ge=-5.0, le=10.0)
    gdp_q4q4: Optional[float] = Field(None, ge=-30.0, le=30.0)
    # Published annual-AVERAGE real GDP growth (MPR Table 2). Reference-only:
    # the engine uses it solely for the annual-average coherence cross-check,
    # never as a model input. Optional so older workbooks still parse.
    gdp_annual_avg: Optional[float] = Field(None, ge=-30.0, le=30.0)
    source_ref: str

    @model_validator(mode="after")
    def _check_range(self) -> "AnnualRow":
        if self.potential_growth_high < self.potential_growth_low:
            raise ValueError(
                f"potential_growth_high ({self.potential_growth_high}) < low "
                f"({self.potential_growth_low}) for year {self.year}"
            )
        if not self.source_ref.strip():
            raise ValueError("source_ref must not be empty")
        return self

    @property
    def potential_growth_mid(self) -> float:
        return (self.potential_growth_low + self.potential_growth_high) / 2.0


class Params(BaseModel):
    """The params sheet, flattened into one validated model.

    All rate-like values are in percent units (2.5 == 2.5%), matching the repo
    convention. Bounds are fail-closed: nonsense transcriptions raise.
    """

    mpr_publication_date: date
    # Last projected quarter the path runs to (e.g. "2028Q4"). Read from the
    # workbook so the horizon travels with the vintage: a later MPR extends the
    # projection table and Jay just punches in the new end quarter. The model,
    # band, validation, and coverage checks all derive from this field — no
    # "2028Q4" literal lives in the engine.
    projection_end_quarter: str
    current_overnight_rate: float = Field(..., ge=-1.0, le=25.0)
    # Output-gap anchor: the BoC staff output-gap estimate at its last published
    # quarter (Valet INDINF_OUTGAPMPR_Q). The model rolls this forward by the gap
    # evolution identity to the seed (MPR) quarter and on through the horizon.
    output_gap_anchor_quarter: str
    output_gap_anchor_value: float = Field(..., ge=-10.0, le=10.0)
    neutral_range_low: float = Field(..., ge=0.0, le=10.0)
    neutral_range_high: float = Field(..., ge=0.0, le=10.0)
    rho: float = Field(0.85, ge=0.0, le=1.0)
    phi_pi: float = Field(4.65, ge=0.0, le=20.0)
    phi_gap: float = Field(0.40, ge=0.0, le=10.0)
    inflation_target: float = Field(2.0, ge=0.0, le=10.0)
    elb_floor: float = Field(0.25, ge=-1.0, le=5.0)
    verified: bool = False

    @field_validator("output_gap_anchor_quarter")
    @classmethod
    def _check_anchor_quarter(cls, v: str) -> str:
        v = str(v).strip()
        if (
            len(v) != 6
            or v[4].upper() != "Q"
            or not v[:4].isdigit()
            or v[5] not in "1234"
        ):
            raise ValueError(
                f"output_gap_anchor_quarter must look like '2025Q4', got {v!r}"
            )
        return v[:4] + "Q" + v[5]

    @field_validator("projection_end_quarter")
    @classmethod
    def _check_projection_end_quarter(cls, v: str) -> str:
        v = str(v).strip()
        if (
            len(v) != 6
            or v[4].upper() != "Q"
            or not v[:4].isdigit()
            or v[5] not in "1234"
        ):
            raise ValueError(
                f"projection_end_quarter must look like '2028Q4', got {v!r}"
            )
        return v[:4] + "Q" + v[5]

    @model_validator(mode="after")
    def _check_ranges(self) -> "Params":
        if self.neutral_range_high < self.neutral_range_low:
            raise ValueError(
                f"neutral_range_high ({self.neutral_range_high}) < "
                f"neutral_range_low ({self.neutral_range_low})"
            )
        # The projection horizon must be after the seed (MPR) quarter; the path
        # iterates forward, so an end at/before the seed is nonsense.
        seed_year = self.mpr_publication_date.year
        seed_qn = (self.mpr_publication_date.month - 1) // 3 + 1
        seed_ord = seed_year * 4 + (seed_qn - 1)
        end_year = int(self.projection_end_quarter[:4])
        end_qn = int(self.projection_end_quarter[5])
        end_ord = end_year * 4 + (end_qn - 1)
        if end_ord <= seed_ord:
            raise ValueError(
                f"projection_end_quarter {self.projection_end_quarter} is not "
                f"after the seed quarter ({seed_year}Q{seed_qn}); the path "
                f"iterates forward from the seed."
            )
        return self

    @property
    def horizon_end_year(self) -> int:
        """Calendar year of projection_end_quarter (coverage checks anchor here)."""
        return int(self.projection_end_quarter[:4])

    @property
    def neutral_nominal_mid(self) -> float:
        return (self.neutral_range_low + self.neutral_range_high) / 2.0


class ShadowInputs(BaseModel):
    """Bundle of all three validated sheets."""

    quarterly: list[QuarterlyRow]
    annual: list[AnnualRow]
    params: Params

    @model_validator(mode="after")
    def _check_nonempty(self) -> "ShadowInputs":
        if not self.quarterly:
            raise ValueError("quarterly sheet is empty")
        if not self.annual:
            raise ValueError("annual sheet is empty")
        # Require at least one q4q4 anchor so the model has something to
        # interpolate toward beyond the near-term quarterly rows.
        if not any(r.anchor_type == "q4q4" for r in self.quarterly):
            raise ValueError("quarterly sheet has no q4q4 anchor rows")
        self._check_no_duplicates()
        return self

    # -- fail-closed integrity checks (audit's reproduced failures) -------- #
    def _check_no_duplicates(self) -> None:
        """Reject duplicate keys that would let a later row silently override.

        Duplicate quarterly quarters, duplicate annual years, and duplicate
        params keys are all transcription mistakes that openpyxl's last-write
        behaviour would otherwise swallow. Params duplicates are caught at
        parse time (a dict), so this covers the two list sheets; params has its
        own guard in ``_parse_params``.
        """
        seen_q: set[str] = set()
        for r in self.quarterly:
            if r.quarter in seen_q:
                raise ValueError(
                    f"duplicate quarterly row for {r.quarter}: a quarter appears "
                    f"twice in the quarterly sheet (the later row would silently "
                    f"override the earlier one)"
                )
            seen_q.add(r.quarter)
        seen_y: set[int] = set()
        for a in self.annual:
            if a.year in seen_y:
                raise ValueError(
                    f"duplicate annual row for year {a.year}: a year appears twice "
                    f"in the annual sheet (the later row would silently override "
                    f"the earlier one)"
                )
            seen_y.add(a.year)

def check_horizon_coverage(inp: "ShadowInputs") -> None:
    """Every horizon year must have Q4 core-CPI coverage and GDP coverage.

    A deleted Q4/Q4 core anchor (e.g. 2026Q4) would otherwise silently produce a
    different interpolated path; a deleted GDP anchor would make
    ``build_gdp_path`` raise mid-run with a less obvious message. We pin the
    coverage requirement up front so the failure names the missing year.

    Enforced at the real workbook entry point (``parse_workbook``), NOT in the
    ShadowInputs validator: builder-level unit tests legitimately construct
    partial bundles to exercise a single helper, and should not be forced to
    carry the full seed-year..horizon anchor set.

    The horizon end is the calendar year of ``projection_end_quarter`` (read from
    the workbook), so coverage requirements travel with the vintage rather than a
    hard-coded year.
    """
    seed_year = inp.params.mpr_publication_date.year
    horizon_end_year = inp.params.horizon_end_year

    # Core CPI: each calendar year from seed_year..horizon_end_year needs at
    # least one core value in Q4 (a quarterly row whose quarter ends in Q4).
    core_q4_years = {
        int(r.quarter[:4])
        for r in inp.quarterly
        if r.quarter[5] == "4"
    }
    for yr in range(seed_year, horizon_end_year + 1):
        if yr not in core_q4_years:
            raise ValueError(
                f"missing core-CPI Q4 coverage for {yr}: every year from the "
                f"seed year ({seed_year}) through {horizon_end_year} (the "
                f"projection_end_quarter year) needs a {yr}Q4 core-CPI value (a "
                f"quarterly anchor or direct quarter); none found"
            )

    # GDP: each horizon year needs either all 4 direct quarters or a q4q4
    # anchor, else the residual fill has nothing to anchor on.
    gdp_direct_by_year: dict[int, set[str]] = {}
    for r in inp.quarterly:
        if r.gdp_growth_qq_ann_forecast is not None:
            yr = int(r.quarter[:4])
            gdp_direct_by_year.setdefault(yr, set()).add(r.quarter[5])
    gdp_anchor_years = {a.year for a in inp.annual if a.gdp_q4q4 is not None}
    for yr in range(seed_year, horizon_end_year + 1):
        has_four_direct = len(gdp_direct_by_year.get(yr, set())) == 4
        if not has_four_direct and yr not in gdp_anchor_years:
            raise ValueError(
                f"missing GDP coverage for {yr}: every horizon year needs "
                f"either 4 direct quarterly GDP values or a Q4/Q4 GDP anchor "
                f"(annual sheet gdp_q4q4); {yr} has neither"
            )


# --------------------------------------------------------------------------- #
# Workbook parsing
# --------------------------------------------------------------------------- #
def _coerce_number(v) -> Optional[float]:
    """Empty cell -> None; otherwise float. Raises on non-numeric junk."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
        return float(v)  # raises ValueError on junk -> fail-closed
    return float(v)


def _coerce_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        s = v.strip().lower()
        if s in ("true", "yes", "y", "1", "verified"):
            return True
        if s in ("false", "no", "n", "0", "unverified", ""):
            return False
    raise ValueError(f"cannot interpret {v!r} as a boolean for 'verified'")


def _header_index(ws, expected: list[str]) -> dict[str, int]:
    """Map column header name -> 0-based column index from the first row."""
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(c).strip() if c is not None else "" for c in header_cells]
    idx = {h: i for i, h in enumerate(headers)}
    missing = [c for c in expected if c not in idx]
    if missing:
        raise ValueError(
            f"sheet {ws.title!r} missing expected columns: {missing}; "
            f"found {headers}"
        )
    return idx


def parse_workbook(path: str | Path) -> ShadowInputs:
    """Parse + validate the punch-in workbook into a ShadowInputs bundle."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"workbook not found: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in ("quarterly", "annual", "params"):
            if sheet not in wb.sheetnames:
                raise ValueError(
                    f"workbook missing required sheet {sheet!r}; "
                    f"found {wb.sheetnames}"
                )

        quarterly = _parse_quarterly(wb["quarterly"])
        annual = _parse_annual(wb["annual"])
        params = _parse_params(wb["params"])
    finally:
        wb.close()

    inp = ShadowInputs(quarterly=quarterly, annual=annual, params=params)
    # Horizon-coverage check fires only at the real workbook entry point.
    check_horizon_coverage(inp)
    return inp


def _parse_quarterly(ws) -> list[QuarterlyRow]:
    cols = [
        "quarter",
        "core_cpi_yoy_forecast",
        "total_cpi_yoy_reference",
        "gdp_growth_qq_ann_forecast",
        "anchor_type",
        "source_ref",
    ]
    idx = _header_index(ws, cols)
    rows: list[QuarterlyRow] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw is None or all(c is None for c in raw):
            continue
        if raw[idx["quarter"]] is None:
            continue
        rows.append(
            QuarterlyRow(
                quarter=str(raw[idx["quarter"]]).strip(),
                core_cpi_yoy_forecast=_coerce_number(raw[idx["core_cpi_yoy_forecast"]]),
                total_cpi_yoy_reference=_coerce_number(raw[idx["total_cpi_yoy_reference"]]),
                gdp_growth_qq_ann_forecast=_coerce_number(raw[idx["gdp_growth_qq_ann_forecast"]]),
                anchor_type=str(raw[idx["anchor_type"]]).strip(),
                source_ref=str(raw[idx["source_ref"]]).strip(),
            )
        )
    return rows


def _parse_annual(ws) -> list[AnnualRow]:
    cols = [
        "year",
        "potential_growth_low",
        "potential_growth_high",
        "gdp_q4q4",
        "source_ref",
    ]
    idx = _header_index(ws, cols)
    # gdp_annual_avg is optional (older workbooks lack it); read it if present.
    avg_idx = idx.get("gdp_annual_avg")
    rows: list[AnnualRow] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw is None or all(c is None for c in raw):
            continue
        if raw[idx["year"]] is None:
            continue
        gdp_avg = (
            _coerce_number(raw[avg_idx])
            if avg_idx is not None and avg_idx < len(raw)
            else None
        )
        rows.append(
            AnnualRow(
                year=int(raw[idx["year"]]),
                potential_growth_low=_coerce_number(raw[idx["potential_growth_low"]]),
                potential_growth_high=_coerce_number(raw[idx["potential_growth_high"]]),
                gdp_q4q4=_coerce_number(raw[idx["gdp_q4q4"]]),
                gdp_annual_avg=gdp_avg,
                source_ref=str(raw[idx["source_ref"]]).strip(),
            )
        )
    return rows


# params keys that are dates / bools / ints / strings get special coercion;
# the rest are floats.
_PARAM_DATE_KEYS = {"mpr_publication_date"}
_PARAM_BOOL_KEYS = {"verified"}
_PARAM_INT_KEYS: set[str] = set()
_PARAM_STR_KEYS = {"output_gap_anchor_quarter", "projection_end_quarter"}

# Deprecated params keys: accepted as no-ops (ignored with a one-line warning) so
# both the original and Jay's cleaned workbooks still parse. The t+4 inflation
# horizon is now fixed by the rule definition (model.RULE_INFLATION_HORIZON_Q),
# not a punch-in field.
_PARAM_DEPRECATED_KEYS = {"inflation_converge_quarters"}


def _parse_params(ws) -> Params:
    idx = _header_index(ws, ["key", "value"])
    kv: dict[str, object] = {}
    seen_keys: set[str] = set()
    # Known params keys; a NOTE/operating row below the block is ignored, but a
    # duplicate of a real key is a fail-closed error (would silently override).
    _known = {
        "mpr_publication_date", "projection_end_quarter",
        "current_overnight_rate",
        "output_gap_anchor_quarter", "output_gap_anchor_value",
        "neutral_range_low", "neutral_range_high", "rho", "phi_pi", "phi_gap",
        "inflation_target", "elb_floor",
        "verified",
    }
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw is None or all(c is None for c in raw):
            continue
        key = raw[idx["key"]]
        if key is None:
            continue
        key = str(key).strip()
        if key in _PARAM_DEPRECATED_KEYS:
            # Deprecated no-op: ignore the value entirely (the t+4 horizon is
            # fixed by the rule definition). Keeps old/cleaned workbooks parseable.
            print(
                f"warning: params key {key!r} is deprecated and ignored; the t+4 "
                f"horizon is fixed by the rule definition"
            )
            continue
        if key in _known:
            if key in seen_keys:
                raise ValueError(
                    f"duplicate params key {key!r}: the key appears twice in the "
                    f"params sheet (the later row would silently override the "
                    f"earlier one)"
                )
            seen_keys.add(key)
        val = raw[idx["value"]]
        if key in _PARAM_DATE_KEYS:
            kv[key] = _coerce_date(val)
        elif key in _PARAM_BOOL_KEYS:
            kv[key] = _coerce_bool(val)
        elif key in _PARAM_INT_KEYS:
            kv[key] = int(_coerce_number(val))
        elif key in _PARAM_STR_KEYS:
            kv[key] = "" if val is None else str(val).strip()
        else:
            kv[key] = _coerce_number(val)
    return Params(**kv)


# Marker written into mpr_publication_date by `make_workbook --new-quarter`. The
# parser rejects it with a clear message so a fresh-quarter workbook cannot be
# run before Jay fills in the real publication date (which would silently reuse
# the prior vintage's seed quarter).
TOFILL_DATE_MARKER = "TO-FILL: MPR publication date (YYYY-MM-DD)"


def _coerce_date(v) -> date:
    if isinstance(v, date):
        return v
    if hasattr(v, "date"):  # datetime
        return v.date()
    if isinstance(v, str):
        s = v.strip()
        if s.upper().startswith("TO-FILL"):
            raise ValueError(
                f"mpr_publication_date is still the TO-FILL placeholder "
                f"({v!r}): a new-quarter workbook was created by "
                f"`make_workbook --new-quarter` but the MPR publication date has "
                f"not been entered yet. Punch in the real date (YYYY-MM-DD) from "
                f"the new MPR before running."
            )
        return date.fromisoformat(s)
    raise ValueError(f"cannot interpret {v!r} as a date for mpr_publication_date")
